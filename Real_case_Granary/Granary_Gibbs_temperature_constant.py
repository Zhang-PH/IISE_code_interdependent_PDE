import numpy as np
from scipy.stats import invgamma, gamma, norm, multivariate_normal
import time


# %%
class GibbsSampler:
    def __init__(self, Y1_obs, B_mat, Beta1, X_d_1, Y_d_1, Z_d_1, T_f_1):
        self.Y1_obs = Y1_obs
        self.B_mat = B_mat
        self.Y1_hat = Y1_obs @ B_mat
        self.Beta1 = Beta1
        self.X_d_1 = X_d_1
        self.Y_d_1 = Y_d_1
        self.Z_d_1 = Z_d_1
        self.T_f_1 = T_f_1
        # self.B_psi = B_psi

        self.psi_11_burn = None
        self.psi_12_burn = None
        self.psi_13_burn = None

        self.psi_11_normal = None
        self.psi_12_normal = None
        self.psi_13_normal = None

        self.beta1_burn = None
        self.beta1_normal = None
        self.gamma_0_burn = None
        self.sigma_e_2_burn = None
        self.gamma_0_normal = None
        self.sigma_e_2_normal = None
        self.nt = None  # grid

    @staticmethod
    def SSE_condition(Y_curr, B_curr, beta_curr):
        return (Y_curr - B_curr @ beta_curr).T @ (Y_curr - B_curr @ beta_curr)

    @staticmethod
    def gamma_post(a_0, b_0, K, zeta_curr):  # gamma posterior
        return gamma(a=a_0 + K / 2, scale=1 / (b_0 + zeta_curr / 2))

    @staticmethod
    def sigma_post(a_e, b_e, n, SSE_curr):  # sigma posterior
        return invgamma(a=a_e + n / 2, scale=b_e + SSE_curr / 2)

    def sigma_2(self, res_sigma, a_e, b_e, n, SSE_new):  # sample sigma_2
        states = []
        sigma_dist = self.sigma_post(a_e, b_e, n, SSE_new)
        cur = res_sigma[-1]  # current

        next1 = sigma_dist.rvs()  # new
        exp1 = sigma_dist.pdf(next1)
        exp2 = sigma_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]

    def gamma_0(self, res_gamma, a_0, b_0, K, zeta_curr):  # sample gamma_0
        states = []
        gamma_dist = self.gamma_post(a_0, b_0, K, zeta_curr)
        cur = res_gamma[-1]  # current
        for i in range(1):
            next1 = gamma_dist.rvs()  # new
            exp1 = gamma_dist.pdf(next1)
            exp2 = gamma_dist.pdf(cur)

            if exp1 >= exp2:
                states.append(next1)
            else:
                states.append(cur)
        return states[-1]

    def F_cal(self, T_f, X_d, Y_d, Z_d, psi_1_curr, psi_2_curr, psi_3_curr):
        return (T_f - psi_1_curr * X_d - psi_2_curr * Y_d - psi_3_curr * Z_d)

    @staticmethod
    def D_cal(B, sigma_curr, gamma_curr, F_curr):
        return (np.linalg.inv(B.T @ B + sigma_curr * gamma_curr * F_curr.T @ F_curr))

    @staticmethod
    def beta_post(D, B, Y, sigma):  # sigma posterior
        return multivariate_normal(mean=D @ B.T @ Y, cov=sigma * D)

    def Beta_sample(self, res_beta, D, B, Y, sigma):  # sample beta
        states = []
        beta_dist = self.beta_post(D, B, Y, sigma)

        cur = res_beta[-1]  # current
        next1 = beta_dist.rvs()  # new
        exp1 = beta_dist.pdf(next1)
        exp2 = beta_dist.pdf(cur)
        if exp2 < exp1:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]

    @staticmethod
    def psi_single(psi_i, sigma_prop):  # psi posterior
        return norm(loc=psi_i, scale=sigma_prop)

    def zeta_cal(self, tm1, tm2, tm3, T_f, X_d, Y_d, Z_d, beta):
        zeta = (T_f - tm1 * X_d - tm2 * Y_d - tm3 * Z_d) @ beta
        return zeta

    def psi_1(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):  # sample psi_1
        states1 = []

        t11_cur = res_t11[-1]  # current
        print('ψ_11:', t11_cur)
        t12_cur = res_t12[-1]
        t13_cur = res_t13[-1]

        new1 = abs(self.psi_single(t11_cur, sigma_prop).rvs())  # new
        next_t11 = new1

        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])
        zeta_new1 = self.zeta_cal(next_t11, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])

        zeta_sumsq_old = zeta_old1.T @ zeta_old1
        zeta_sumsq_new = zeta_new1.T @ zeta_new1

        square_cur = (t11_cur ** 2 + t12_cur ** 2 + t13_cur ** 2) / (2 * sigma_psi)
        square_new = (next_t11 ** 2 + t12_cur ** 2 + t13_cur ** 2) / (2 * sigma_psi)

        exp1 = (-square_cur - gamma_0 * zeta_sumsq_old / 2)
        exp2 = (-square_new - gamma_0 * zeta_sumsq_new / 2)

        if exp2 > exp1:
            states1.append(next_t11)
            # t11_cur = next_t11
        else:
            states1.append(t11_cur)
        print('zeta:', zeta_sumsq_old, 'zeta new:', zeta_sumsq_new)
        return states1[-1]

    def psi_2(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):  # sample psi_2
        states2 = []

        t11_cur = res_t11[-1]
        t12_cur = res_t12[-1]  # current
        t13_cur = res_t13[-1]
        print('ψ_12:', t12_cur)

        new1 = abs(self.psi_single(t12_cur, sigma_prop).rvs())  # new
        next_t12 = new1

        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])
        zeta_new1 = self.zeta_cal(t11_cur, next_t12, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])

        zeta_sumsq_old = zeta_old1.T @ zeta_old1
        zeta_sumsq_new = zeta_new1.T @ zeta_new1

        square_cur = (t11_cur ** 2 + t12_cur ** 2 + t13_cur ** 2) / (2 * sigma_psi)
        square_new = (t11_cur ** 2 + next_t12 ** 2 + t13_cur ** 2) / (2 * sigma_psi)

        exp1 = (-square_cur - gamma_0 * zeta_sumsq_old / 2)
        exp2 = (-square_new - gamma_0 * zeta_sumsq_new / 2)
        if exp2 > exp1:
            states2.append(next_t12)
        else:
            states2.append(t12_cur)
        print('zeta:', zeta_sumsq_old, 'zeta new:', zeta_sumsq_new)
        return states2[-1]

    def psi_3(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):  # sample psi_3
        states3 = []

        t11_cur = res_t11[-1]
        t12_cur = res_t12[-1]
        t13_cur = res_t13[-1]  # current
        print('ψ_13:', t13_cur)

        new1 = abs(self.psi_single(t13_cur, sigma_prop).rvs())  # new
        next_t13 = new1

        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])
        zeta_new1 = self.zeta_cal(t11_cur, t12_cur, next_t13, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])

        zeta_sumsq_old = zeta_old1.T @ zeta_old1
        zeta_sumsq_new = zeta_new1.T @ zeta_new1

        square_cur = (t11_cur ** 2 + t12_cur ** 2 + t13_cur ** 2) / (2 * sigma_psi)
        square_new = (t11_cur ** 2 + t12_cur ** 2 + next_t13 ** 2) / (2 * sigma_psi)

        exp1 = (-square_cur - gamma_0 * zeta_sumsq_old / 2)
        exp2 = (-square_new - gamma_0 * zeta_sumsq_new / 2)
        if exp2 > exp1:
            states3.append(next_t13)
        else:
            states3.append(t13_cur)
        print('zeta:', zeta_sumsq_old, 'zeta new:', zeta_sumsq_new)
        return states3[-1]

    def gibbs_burn(self, N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt, shape):
        res_sigma = []
        res_gamma = []
        res_psi11 = []  # PDE1 2nd derivative
        res_psi12 = []  # PDE1 1st derivative
        res_psi13 = []  # PDE constant
        res_beta1 = []
        self.nt = nt
        zeta_burn = []

        # 1. Bspline error
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.Beta1)
        SSE_new = SSE_new1

        # 2. time-varying parameters initialization
        psi_11_curr = abs(np.random.normal(0, sigma_psi))
        psi_12_curr = abs(np.random.normal(0, sigma_psi))
        psi_13_curr = abs(np.random.normal(0, sigma_psi))

        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)

        # 3. sigma initialization
        res_sigma.append((self.sigma_post(a_e, b_e, n, SSE_new)).rvs())

        # 4. beta initialization
        res_beta1.append(self.Beta1)

        # 5. PDE error  initialization
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  self.Beta1)  # zeta initialization
        zeta_new = zeta_new1.T @ zeta_new1  # PDE  error
        print('zeta origin', zeta_new)
        zeta_burn.append(zeta_new)

        # 6. gamma initialization
        res_gamma.append(self.gamma_post(a_0, b_0, K, zeta_new).rvs())
        # iteration
        for i in range(N_burn):
            # 7. update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))

            # 8. update gamma
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma[-1])

            if i <= 100:
                F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1, res_psi11[-1], res_psi12[-1],
                                    res_psi13[-1])
                D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
                # 9. update beta
                res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            else:
                res_beta1.append(res_beta1[-1])

            # 10. update time-varying parameter ψ
            sample11 = self.psi_1(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            sample12 = self.psi_2(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            sample13 = self.psi_3(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)

            # 11. update Bspline error
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_new = SSE_1

            # 12. update PDE error
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], self.T_f_1, self.X_d_1, self.Y_d_1,
                                   self.Z_d_1, res_beta1[-1])
            zeta_new = zeta_1.T @ zeta_1
            zeta_burn.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new)
            print('------------------')

        self.sigma_e_2_burn = res_sigma[-1]
        self.gamma_0_burn = res_gamma[-1]
        self.beta1_burn = res_beta1[-1]
        self.psi_11_burn = res_psi11[-1]
        self.psi_12_burn = res_psi12[-1]
        self.psi_13_burn = res_psi13[-1]

        print("sigma_e_2 estimation:", res_sigma[-1])
        print("gamma_0 estimation:", res_gamma[-1])
        print("beta1 estimation:", res_beta1[-1])

        print("psi_1 estimation:", res_psi11[-1])
        print("psi_2 estimation:", res_psi12[-1])
        print("psi_3 estimation:", res_psi13[-1])
        return res_psi11, res_psi12, res_psi13, res_beta1, res_sigma, res_gamma, zeta_burn

    def gibbs_normal(self, N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt_0):
        res_sigma = []
        res_gamma = []
        res_psi11 = []  # PDE1 2nd derivative
        res_psi12 = []  # PDE1 1st derivative
        res_psi13 = []  # PDE constant
        res_beta1 = []
        self.nt = nt_0
        zeta_normal = []  # zeta²

        # 1. Bspline error  initialization
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.beta1_burn)
        SSE_new = SSE_new1

        # 2. time-varying parameters initialization
        psi_11_curr = self.psi_11_burn
        psi_12_curr = self.psi_12_burn
        psi_13_curr = self.psi_13_burn

        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)

        # 3. sigma initialization
        res_sigma.append(self.sigma_e_2_burn)

        # 4. beta  initialization
        res_beta1.append(self.beta1_burn)

        # 5. PDE error  initialization
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])  # zeta initialization
        zeta_new = zeta_new1.T @ zeta_new1
        zeta_normal.append(zeta_new)

        # 6. gamma  initialization
        res_gamma.append(self.gamma_0_burn)  # gamma initialization
        # iteration
        for i in range(N_normal):
            # 7. update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))

            # 8. update gamma
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma[-1])

            F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1, res_psi11[-1], res_psi12[-1],
                                res_psi13[-1])
            D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
            # 9.  update beta
            # res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            res_beta1.append(res_beta1[-1])

            # 10. update time-varying parameter ψ
            sample11 = self.psi_1(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            sample12 = self.psi_2(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            sample13 = self.psi_3(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)

            # 11. update Bspline error
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_new = SSE_1

            # 12. update PDE error
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], self.T_f_1, self.X_d_1, self.Y_d_1,
                                   self.Z_d_1, res_beta1[-1])
            zeta_new = zeta_1.T @ zeta_1
            zeta_normal.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new)
            print('------------------')

        self.sigma_e_2_normal = res_sigma[-1]
        self.gamma_0_normal = res_gamma[-1]
        self.beta1_normal = res_beta1[-1]
        self.psi_11_normal = res_psi11[-1]
        self.psi_12_normal = res_psi12[-1]
        self.psi_13_normal = res_psi13[-1]

        print("sigma_e_2 estimation:", res_sigma[-1])
        print("gamma_0 estimation:", res_gamma[-1])
        print("beta1 estimation:", res_beta1[-1])

        print("psi_1 estimation:", res_psi11[-1])
        print("psi_2 estimation:", res_psi12[-1])
        print("psi_3 estimation:", res_psi13[-1])
        return res_psi11, res_psi12, res_psi13, res_beta1, res_sigma, res_gamma, zeta_normal


# %%
# true value
B_psi = np.load('input/B_psi_91.npy')

Y_t = np.load('input/Y_temp.npy')
B_t = np.load('input/B_mat_temp.npy')

beta_ini_t = np.load('input/beta_temp.npy')
Y_hat_t = B_t @ beta_ini_t

X_d_t = np.load('input/X_d_temp.npy')
Y_d_t = np.load('input/Y_d_temp.npy')
Z_d_t = np.load('input/Z_d_temp.npy')
T_f_t = np.load('input/T_f_temp.npy')
T_f_t = T_f_t

n = B_t.shape[0]  # observation
K = beta_ini_t.shape[0]  # number of basis functions
a_e = 0.001
b_e = 0.001
a_0 = 0.001
b_0 = 0.001
sigma_psi = 1

sigma_prop1 = 0.01
sigma_prop2 = 0.0001

nt = 54
shape_psi = B_psi.shape[1]
N_burn = 2000
N_normal = 10000
# %%
start_time11 = time.time()
GibbsObj_1 = GibbsSampler(Y_t, B_t, beta_ini_t, X_d_t, Y_d_t, Z_d_t, T_f_t)
res_psi11, res_psi12, res_psi13, res_beta1, res_sigma1, res_gamma1, zeta_burn = GibbsObj_1.gibbs_burn(N_burn, n, K, a_e,
                                                                                                      b_e, a_0, b_0,
                                                                                                      sigma_prop1, nt,
                                                                                                      shape_psi)
end_time11 = time.time()
print("PDE temperature burn_in stage running time:", end_time11 - start_time11, "s")
# %%
# PDE1 steady stage
start_time12 = time.time()
res_psi11_normal, res_psi12_normal, res_psi13_normal, res_beta1_normal, res_sigma1_normal, res_gamma1_normal, zeta_normal = GibbsObj_1.gibbs_normal(
    N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop2, nt)
end_time12 = time.time()
print("PDE temperature steady_stage running time:", end_time12 - start_time12, "s")
# %%
psi_11_hat = res_psi11_normal[-1]
psi_12_hat = res_psi12_normal[-1]
psi_13_hat = res_psi13_normal[-1]

beta1_hat = res_beta1_normal[-1]

sigma_hat = res_sigma1_normal[-1]
gamma_hat = res_gamma1_normal[-1]
# %%
# PDE error
zeta_ob1 = GibbsObj_1.zeta_cal(psi_11_hat, psi_12_hat, psi_13_hat, T_f_t, X_d_t, Y_d_t, Z_d_t, beta1_hat)

zeta_ob_ss = zeta_ob1.T @ zeta_ob1

print('mean_zeta', np.mean(zeta_ob1))
print('mean_abs_zeta', np.mean(np.abs(zeta_ob1)))
print('max_zeta', np.max(zeta_ob1))
print('min_zeta', np.min(zeta_ob1))
print('rmse_zeta', np.sqrt(zeta_ob_ss / zeta_ob1.shape[0]))
print('SST', zeta_ob_ss)
# %%
# Bspline error
error_bs1 = Y_t - B_t @ beta1_hat

error_ss = error_bs1.T @ error_bs1

print('mean_BSpline', np.mean(error_bs1))
print('mean_abs_BSpline', np.mean(np.abs(error_bs1)))
print('max_BSpline', np.max(error_bs1))
print('min_BSpline', np.min(error_bs1))
print('rmse_BSpline', np.sqrt(error_ss / error_bs1.shape[0]))
print('SST_BSpline', error_ss)
# %%
np.save('output_temp_constant/psi_1_burn.npy', res_psi11)
np.save('output_temp_constant/psi_2_burn.npy', res_psi12)
np.save('output_temp_constant/psi_3_burn.npy', res_psi13)
np.save('output_temp_constant/beta_burn.npy', res_beta1)

np.save('output_temp_constant/sigma_burn.npy', res_sigma1)
np.save('output_temp_constant/gamma_burn.npy', res_gamma1)

np.save('output_temp_constant/psi_1_steady.npy', res_psi11_normal)
np.save('output_temp_constant/psi_2_steady.npy', res_psi12_normal)
np.save('output_temp_constant/psi_3_steady.npy', res_psi13_normal)
np.save('output_temp_constant/beta_steady.npy', res_beta1_normal)

np.save('output_temp_constant/sigma_steady.npy', res_sigma1_normal)
np.save('output_temp_constant/gamma_steady.npy', res_gamma1_normal)
# %%
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import datetime

now = datetime.datetime.now()

wb = Workbook()
dest_filename = "%s_Temperature_Constant_result.xlsx" % (now.strftime('%Y%m%d-%H%M%S'))

ws = wb.active
ws.title = 'Result'

ws['A1'] = "Date"
ws['B1'] = now.strftime('%Y-%m-%d')  # %H:%M:%S
ws['C1'] = "Result"
ws.merge_cells(range_string='C1:E1')

ws.append(["psi1", '采样估计值'] + [psi_11_hat for i in range(shape_psi)])
ws.append(["psi2", '采样估计值'] + [psi_12_hat for i in range(shape_psi)])
ws.append(["psi3", '采样估计值'] + [psi_13_hat for i in range(shape_psi)])

ws.append(["sigma_2", '采样估计值', sigma_hat, (b_e + error_bs1.T @ error_bs1 / 2) / (a_e + n / 2 - 1)])
ws.append(["gamma", '采样估计值', gamma_hat, (a_0 + K / 2) / (b_0 + zeta_ob1.T @ zeta_ob1 / 2)])

ws['I1'] = "Model Errors"
ws.merge_cells(range_string='I1:L1')

ws['I2'] = "PDE temperature"
ws.merge_cells(range_string='I2:I7')
# ws['I14'] = "PDE humidity"
# ws.merge_cells(range_string='I14:I19')

ws['I8'] = "统计模型 temperature"
ws.merge_cells(range_string='I8:I13')
# ws['I20'] = "统计模型 humidity"
# ws.merge_cells(range_string='I20:I25')


J = ['mean_zeta', 'mean_abs_zeta', 'max_zeta', 'min_zeta', 'rmse_zeta', 'SST_zeta', 'mean_BSpline', 'mean_abs_BSpline',
     'max_BSpline', 'min_BSpline', 'rmse_BSpline', 'SST_BSpline']
K1 = ['均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和', '均值', '绝对值均值', '最大值', '最小值',
      '均方根误差', '总误差平方和']

L1 = [np.mean(zeta_ob1), np.mean(np.abs(zeta_ob1)), np.max(zeta_ob1), np.min(zeta_ob1),
      np.sqrt(zeta_ob1.T @ zeta_ob1 / zeta_ob1.shape[0]), zeta_ob1.T @ zeta_ob1, np.mean(error_bs1),
      np.mean(np.abs(error_bs1)), np.max(error_bs1), np.min(error_bs1),
      np.sqrt(error_bs1.T @ error_bs1 / error_bs1.shape[0]), error_bs1.T @ error_bs1]
# L2 = [np.mean(zeta_ob2), np.mean(np.abs(zeta_ob2)), np.max(zeta_ob2), np.min(zeta_ob2), np.sqrt(zeta_ob2.T@zeta_ob2/zeta_ob2.shape[0]), zeta_ob2.T@zeta_ob2,np.mean(error_bs2), np.mean(np.abs(error_bs2)), np.max(error_bs2), np.min(error_bs2), np.sqrt(error_bs2.T@error_bs2/error_bs2.shape[0]), error_bs2.T@error_bs2]

for i in range(2, 14):
    ws['J%d' % i] = J[i - 2]
    ws['K%d' % i] = K1[i - 2]
    ws['L%d' % i] = L1[i - 2]

    # ws['J%d'%(i+12)] = J[i-2]
    # ws['K%d'%(i+12)] = K1[i-2]
    # ws['L%d'%(i+12)] = L2[i-2]

# ws['K26'] = "PDE_SST"
# ws['K27'] = "B_Spline_SST"
#
# ws['L26'] = zeta_ob1.T@zeta_ob1 + zeta_ob2.T@zeta_ob2
# ws['L27'] = error_bs1.T@error_bs1 + error_bs2.T@error_bs2

N = ['#', 'Ver.', 'p_Bspline', 'p_Bpsi', 'n', 'K', 'a_e', 'b_e', 'a_0', 'b_0', 'sigma_psi', 'sigma_prop1',
     'sigma_prop2', 'nt', 'N_burn', 'N_normal']
O = ['Value', 'Temperature_Constant', '3', '3', n, K, a_e, b_e, a_0, b_0, sigma_psi, sigma_prop1, sigma_prop2, nt,
     N_burn, N_normal]
P = ['含义', '版本', '统计建模样条阶数（次数+1）', '时变参数样条阶数（次数+1）', '观测值个数', '基函数个数', '', '', '', '',
     '', '第一阶段采样sigma（Burn-in stage）', '第二阶段采样sigma（Steady stage）', '', 'Burn-in 阶段采样次数',
     'Steady阶段采样次数']

for i in range(1, 17):
    ws['N%d' % i] = N[i - 1]
    ws['O%d' % i] = O[i - 1]
    ws['P%d' % i] = P[i - 1]


def setCellStyle(st):
    """
    设置单元格格式
    params: st:就是第一步创建的工作表
    """

    # 边框
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'),
    )

    # 对齐
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        indent=0
    )

    # 字体
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        strike=False,
        color='000000'
    )

    for row, row_ind in zip(st.iter_rows(), range(1, st.max_row + 1)):
        for cell in row:
            # 设置边框
            st[cell.coordinate].border = border
            # 设置居中对齐
            st[cell.coordinate].alignment = alignment
            # 行高40
            st.row_dimensions[row_ind].height = 14
            # 设置字体
            st[cell.coordinate].font = font

    # 设置列宽
    st.column_dimensions['A'].width = 6
    st.column_dimensions['B'].width = 14
    st.column_dimensions['C'].width = 14
    st.column_dimensions['D'].width = 14
    st.column_dimensions['E'].width = 14
    st.column_dimensions['F'].width = 14
    st.column_dimensions['G'].width = 14

    st.column_dimensions['I'].width = 14
    st.column_dimensions['J'].width = 14
    st.column_dimensions['k'].width = 14
    st.column_dimensions['L'].width = 14
    st.column_dimensions['N'].width = 14
    st.column_dimensions['O'].width = 35
    st.column_dimensions['P'].width = 35

    st.column_dimensions['H'].width = 2
    st.column_dimensions['M'].width = 2


setCellStyle(ws)
wb.save(dest_filename)
wb.close()