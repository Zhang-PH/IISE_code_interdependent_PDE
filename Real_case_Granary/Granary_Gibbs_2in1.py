import numpy as np
from scipy.stats import invgamma, gamma, norm, multivariate_normal, pearsonr
import copy
import time


# %%
class GibbsSampler:
    def __init__(self, Y1_obs, Y2_obs, B_mat, Beta1, Beta2, X_d_1, Y_d_1, Z_d_1, T_f_1, X_d_2, Y_d_2, Z_d_2, T_f_2,
                 B_psi):
        self.Y1_obs = Y1_obs
        self.Y2_obs = Y2_obs
        self.B_mat = B_mat
        self.Y1_hat = Y1_obs @ B_mat
        self.Y2_hat = Y2_obs @ B_mat
        self.Beta1 = Beta1
        self.Beta2 = Beta2
        self.X_d_1 = X_d_1
        self.Y_d_1 = Y_d_1
        self.Z_d_1 = Z_d_1
        self.T_f_1 = T_f_1
        self.X_d_2 = X_d_2
        self.Y_d_2 = Y_d_2
        self.Z_d_2 = Z_d_2
        self.T_f_2 = T_f_2
        self.B_psi = B_psi

        self.psi_11_burn = None
        self.psi_12_burn = None
        self.psi_13_burn = None

        self.psi_21_burn = None
        self.psi_22_burn = None
        self.psi_23_burn = None

        self.psi_11_normal = None
        self.psi_12_normal = None
        self.psi_12_normal = None
        self.psi_13_normal = None

        self.psi_21_normal = None
        self.psi_22_normal = None
        self.psi_23_normal = None

        self.beta1_burn = None
        self.beta2_burn = None
        self.beta1_normal = None
        self.beta2_normal = None
        self.gamma_0_burn = None
        self.sigma_e_2_burn = None
        self.gamma_0_normal = None
        self.sigma_e_2_normal = None
        self.nt = None  # grid
        self.psi11_set = []
        self.psi12_set = []
        self.psi13_set = []
        self.psi21_set = []
        self.psi22_set = []
        self.psi23_set = []
        self.sigma_set = []
        self.gamma_set = []
        self.beta1_set = []
        self.beta2_set = []

    @staticmethod
    def SSE_condition(Y_curr, B_curr, beta_curr):
        return (Y_curr - B_curr @ beta_curr).T @ (Y_curr - B_curr @ beta_curr)

    def psi_mat(self, t_cur):
        temp = self.B_psi @ t_cur
        # con = temp
        # for i in range(self.nt - 1):
        #     con = np.concatenate((con, temp),axis=0)
        con = np.array([temp for i in range(self.nt)])
        return con.reshape(self.nt * (temp.shape[0]))

    @staticmethod
    def gamma_post(a_0, b_0, K, zeta_curr):  # gamma posterior
        return gamma(a=a_0 + K / 2, scale=1 / (b_0 + zeta_curr / (2 * 2)))

    @staticmethod
    def sigma_post(a_e, b_e, n, SSE_curr):  # sigma posterior
        return invgamma(a=a_e + n / 2, scale=b_e + SSE_curr / (2 * 2))

    def sigma_2(self, res_sigma, a_e, b_e, n, SSE_new):  # sampling sigma_2
        states = []
        sigma_dist = self.sigma_post(a_e, b_e, n, SSE_new)
        cur = res_sigma[-1]  # current

        next1 = sigma_dist.rvs()  # new
        (self.sigma_set).append(next1)
        exp1 = sigma_dist.pdf(next1)
        exp2 = sigma_dist.pdf(cur)

        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]

    def gamma_0(self, res_gamma, a_0, b_0, K, zeta_curr):  # sampling gamma_0
        states = []
        gamma_dist = self.gamma_post(a_0, b_0, K, zeta_curr)
        cur = res_gamma[-1]  # current

        next1 = gamma_dist.rvs()  # new
        (self.gamma_set).append(next1)
        exp1 = gamma_dist.pdf(next1)
        exp2 = gamma_dist.pdf(cur)

        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]

    def F_cal(self, T_f, X_d, Y_d, Z_d, psi_1_curr, psi_2_curr, psi_3_curr):
        return (T_f - np.diag(self.psi_mat(psi_1_curr)) @ X_d - np.diag(self.psi_mat(psi_2_curr)) @ Y_d - np.diag(
            self.psi_mat(psi_3_curr)) @ Z_d)

    @staticmethod
    def D_cal(B, sigma_curr, gamma_curr, F_curr):
        return (np.linalg.inv(B.T @ B + sigma_curr * gamma_curr * F_curr.T @ F_curr))

    @staticmethod
    def beta_post(D, B, Y, sigma):  # sigma posterior
        return multivariate_normal(mean=D @ B.T @ Y, cov=sigma * D)

    def Beta_sample(self, res_beta, D, B, Y, sigma):  # sampling beta
        beta_dist = self.beta_post(D, B, Y, sigma)
        cur = res_beta[-1]  # current
        next1 = beta_dist.rvs()  # new
        exp1 = beta_dist.logpdf(next1)
        exp2 = beta_dist.logpdf(cur)
        if exp2 < exp1:
            return next1
        else:
            return cur

    @staticmethod
    def psi_single(psi_i, sigma_prop):  # psi posterior
        return norm(loc=psi_i, scale=sigma_prop)

    def zeta_cal(self, tm1, tm2, tm3, T_f, X_d, Y_d, Z_d, beta):
        tm1_mat = self.psi_mat(tm1)
        tm2_mat = self.psi_mat(tm2)
        tm3_mat = self.psi_mat(tm3)
        zeta = T_f @ beta - np.multiply(tm1_mat, X_d @ beta) - np.multiply(tm2_mat, Y_d @ beta) - np.multiply(tm3_mat,
                                                                                                              Z_d @ beta)
        return zeta

    @staticmethod
    def p_cal(a, b):
        pp = 1 - 0.01 * (pearsonr(a, b)[0])
        return pp

    def psi_1(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_beta1, res_beta2, sigma_prop,
              gamma_0):  # sampling psi_x1
        states1 = []
        states2 = []

        t11_cur = res_t11  # current
        print('ψ_11:', t11_cur)
        t12_cur = res_t12
        t13_cur = res_t13

        t21_cur = res_t21
        print('ψ_21:', t21_cur)
        t22_cur = res_t22
        t23_cur = res_t23

        SS = np.array([[1, 0.9], [0.9, 1]])
        INV_SIGMA = np.linalg.pinv(SS)  # new
        tx2 = np.array([t12_cur, t22_cur])  # new
        tx3 = np.array([t13_cur, t23_cur])  # new

        shape = len(t11_cur)
        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        new_psi1 = []
        new_psi2 = []
        for i in range(shape):  # t11
            new1 = abs(self.psi_single(t11_cur[i], sigma_prop).rvs())  # new
            next_t11 = copy.deepcopy(t11_cur)
            next_t11[i] = new1
            new_psi1.append(new1)

            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)

            zeta_new1 = self.zeta_cal(next_t11, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)

            p12old = self.p_cal(t11_cur, t21_cur)
            p12new = self.p_cal(next_t11, t21_cur)
            # print('(p12old:',p12old, 'p12new:', p12new)

            zeta_sumsq_old1.append(zeta_old1.T @ zeta_old1 + p12old * zeta_old2.T @ zeta_old2)
            zeta_sumsq_new1.append(zeta_new1.T @ zeta_new1 + p12new * zeta_old2.T @ zeta_old2)
            # print(zeta_sumsq_old1, zeta_sumsq_new1[-1])
            tx1 = np.array([t11_cur, t21_cur])
            tx1_new = np.array([next_t11, t21_cur])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1_new.T @ INV_SIGMA @ tx1_new) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old1[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new1[-1] / 2)

            if exp2 > exp1:
                states1.append(next_t11)
                t11_cur = next_t11
            else:
                states1.append(t11_cur)
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []

        for i in range(shape):  # t21
            new2 = abs(self.psi_single(t21_cur[i], sigma_prop).rvs())  # new
            next_t21 = copy.deepcopy(t21_cur)
            next_t21[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)

            zeta_new2 = self.zeta_cal(next_t21, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)

            p21old = self.p_cal(t21_cur, t11_cur)
            p21new = self.p_cal(next_t21, t11_cur)

            # print('(p21old:',p21old, 'p21new:', p21new)
            zeta_sumsq_old2.append(p21old * (zeta_old1.T @ zeta_old1) + zeta_old2.T @ zeta_old2)
            zeta_sumsq_new2.append(p21new * (zeta_old1.T @ zeta_old1) + zeta_new2.T @ zeta_new2)
            # print(zeta_sumsq_old2[-1], zeta_sumsq_new2[-1])

            tx1 = np.array([t11_cur, t21_cur])
            tx1_new = np.array([t11_cur, next_t21])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1_new.T @ INV_SIGMA @ tx1_new) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)

            # exp1 = np.exp(-square_cur/2 - gamma_0*zeta_sumsq_old2[-1]/2)
            # exp2 = np.exp(-square_new/2 - gamma_0*zeta_sumsq_new2[-1]/2)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old2[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new2[-1] / 2)

            if exp2 > exp1:
                states2.append(next_t21)
                t21_cur = next_t21
            else:
                states2.append(t21_cur)

        # print('zeta:',zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2)
        self.psi11_set.append(new_psi1)
        self.psi21_set.append(new_psi2)
        print('p*zeta:', zeta_sumsq_old2[-1], 'p*zeta new:', zeta_sumsq_new2[-1])
        return states1[-1], states2[-1]

    def psi_2(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_beta1, res_beta2, sigma_prop,
              gamma_0):  # sampling psi_x2
        states1 = []
        states2 = []

        t11_cur = res_t11
        t12_cur = res_t12  # current
        t13_cur = res_t13
        print('ψ_12:', t12_cur)

        t21_cur = res_t21
        t22_cur = res_t22
        print('ψ_22:', t22_cur)
        t23_cur = res_t23

        SS = np.array([[1, 0.9], [0.9, 1]])
        INV_SIGMA = np.linalg.pinv(SS)  # new
        tx1 = np.array([t11_cur, t21_cur])  # new
        tx3 = np.array([t13_cur, t23_cur])  # new

        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        new_psi1 = []
        new_psi2 = []
        shape = len(t12_cur)
        for i in range(shape):  # t12
            new1 = abs(self.psi_single(t12_cur[i], sigma_prop).rvs())  # new
            next_t12 = copy.deepcopy(t12_cur)
            next_t12[i] = new1
            new_psi1.append(new1)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)
            zeta_new1 = self.zeta_cal(t11_cur, next_t12, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)

            p12old = self.p_cal(t12_cur, t22_cur)
            p12new = self.p_cal(next_t12, t22_cur)
            # print('(p12old:',p12old, 'p12new:', p12new)

            zeta_sumsq_old1.append(zeta_old1.T @ zeta_old1 + p12old * zeta_old2.T @ zeta_old2)
            zeta_sumsq_new1.append(zeta_new1.T @ zeta_new1 + p12new * zeta_old2.T @ zeta_old2)
            # print(zeta_sumsq_old1[-1], zeta_sumsq_new1[-1])

            tx2 = np.array([t12_cur, t22_cur])
            tx2_new = np.array([next_t12, t22_cur])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2_new.T @ INV_SIGMA @ tx2_new) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old1[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new1[-1] / 2)
            if exp2 > exp1:
                states1.append(next_t12)
                t12_cur = next_t12
            else:
                states1.append(t12_cur)
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []

        for i in range(shape):  # t22
            new2 = abs(self.psi_single(t22_cur[i], sigma_prop).rvs())  # new
            next_t22 = copy.deepcopy(t22_cur)
            next_t22[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)
            zeta_new2 = self.zeta_cal(t21_cur, next_t22, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)

            p21old = self.p_cal(t22_cur, t12_cur)
            p21new = self.p_cal(next_t22, t12_cur)
            # print('(p21old:',p21old, 'p21new:', p21new)

            zeta_sumsq_old2.append(p21old * (zeta_old1.T @ zeta_old1) + zeta_old2.T @ zeta_old2)
            zeta_sumsq_new2.append(p21new * (zeta_old1.T @ zeta_old1) + zeta_new2.T @ zeta_new2)
            # print(zeta_sumsq_old2[-1], zeta_sumsq_new2[-1])

            tx2 = np.array([t12_cur, t22_cur])
            tx2_new = np.array([t12_cur, next_t22])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2_new.T @ INV_SIGMA @ tx2_new) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old2[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new2[-1] / 2)
            if exp2 > exp1:
                states2.append(next_t22)
                t22_cur = next_t22
            else:
                states2.append(t22_cur)
        # print('zeta:',zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2)
        self.psi12_set.append(new_psi1)
        self.psi22_set.append(new_psi2)
        print('p*zeta:', zeta_sumsq_old2[-1], 'p*zeta new:', zeta_sumsq_new2[-1])
        return states1[-1], states2[-1]

    def psi_3(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_beta1, res_beta2, sigma_prop,
              gamma_0):  # sampling psi_x3
        states1 = []
        states2 = []

        t11_cur = res_t11
        t12_cur = res_t12
        t13_cur = res_t13  # current
        print('ψ_13:', t13_cur)
        t21_cur = res_t21
        t22_cur = res_t22
        t23_cur = res_t23
        print('ψ_23:', t23_cur)

        SS = np.array([[1, 0.9], [0.9, 1]])
        INV_SIGMA = np.linalg.pinv(SS)  # new

        tx1 = np.array([t11_cur, t21_cur])  # new
        tx2 = np.array([t12_cur, t22_cur])  # new

        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        shape = len(t13_cur)
        new_psi1 = []
        new_psi2 = []
        for i in range(shape):  # t13
            new1 = abs(self.psi_single(t13_cur[i], sigma_prop).rvs())  # new
            next_t13 = copy.deepcopy(t13_cur)
            next_t13[i] = new1
            new_psi1.append(new1)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)
            zeta_new1 = self.zeta_cal(t11_cur, t12_cur, next_t13, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)

            p12old = self.p_cal(t13_cur, t23_cur)
            p12new = self.p_cal(next_t13, t23_cur)
            # print('(p12old:',p12old, 'p12new:', p12new)

            zeta_sumsq_old1.append(zeta_old1.T @ zeta_old1 + p12old * zeta_old2.T @ zeta_old2)
            zeta_sumsq_new1.append(zeta_new1.T @ zeta_new1 + p12new * zeta_old2.T @ zeta_old2)
            # print(zeta_sumsq_old1[-1], zeta_sumsq_new1[-1])

            tx3 = np.array([t13_cur, t23_cur])  # new
            tx3_new = np.array([next_t13, t23_cur])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3_new.T @ INV_SIGMA @ tx3_new)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old1[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new1[-1] / 2)
            if exp2 > exp1:
                states1.append(next_t13)
                t13_cur = next_t13
            else:
                states1.append(t13_cur)
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []

        for i in range(shape):  # t23
            new2 = abs(self.psi_single(t23_cur[i], sigma_prop).rvs())  # new
            next_t23 = copy.deepcopy(t23_cur)
            next_t23[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                      res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)
            zeta_new2 = self.zeta_cal(t21_cur, t22_cur, next_t23, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                      res_beta2)

            p21old = self.p_cal(t23_cur, t13_cur)
            p21new = self.p_cal(next_t23, t13_cur)
            # print('(p21old:',p21old, 'p21new:', p21new)

            zeta_sumsq_old2.append(p21old * (zeta_old1.T @ zeta_old1) + zeta_old2.T @ zeta_old2)
            zeta_sumsq_new2.append(p21new * (zeta_old1.T @ zeta_old1) + zeta_new2.T @ zeta_new2)
            # print(zeta_sumsq_old2[-1], zeta_sumsq_new2[-1])

            tx3 = np.array([t13_cur, t23_cur])  # new
            tx3_new = np.array([t13_cur, next_t23])

            square_cur = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3.T @ INV_SIGMA @ tx3)
            square_new = np.trace(tx1.T @ INV_SIGMA @ tx1) + np.trace(tx2.T @ INV_SIGMA @ tx2) + np.trace(
                tx3_new.T @ INV_SIGMA @ tx3_new)

            exp1 = (-square_cur / 2 - gamma_0 * zeta_sumsq_old2[-1] / 2)
            exp2 = (-square_new / 2 - gamma_0 * zeta_sumsq_new2[-1] / 2)
            if exp2 > exp1:
                states2.append(next_t23)
                t23_cur = next_t23
            else:
                states2.append(t23_cur)
        # print('zeta:',zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2)
        print('p*zeta:', zeta_sumsq_old2[-1], 'p*zeta new:', zeta_sumsq_new2[-1])
        # print('exp1', exp1)
        # print('exp2', exp2)
        self.psi13_set.append(new_psi1)
        self.psi23_set.append(new_psi2)
        return states1[-1], states2[-1]

    def gibbs_burn(self, N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt, shape):
        res_sigma = []
        res_gamma = []
        res_psi11 = []  # PDE1 2nd derivative
        res_psi12 = []  # PDE1 1st derivative
        res_psi13 = []  # PDE1 constant
        res_psi21 = []  # PDE2 2nd derivative
        res_psi22 = []  # PDE2 1st derivative
        res_psi23 = []  # PDE2 constant

        res_beta1 = []
        res_beta2 = []
        self.nt = nt
        zeta_burn = []

        # 1.Bspline error
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.Beta1)
        SSE_new2 = self.SSE_condition(self.Y2_obs, self.B_mat, self.Beta2)
        SSE_new = SSE_new1 + SSE_new2

        # 2.time-varying parameters initialization
        psi_11_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        psi_12_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        psi_13_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        psi_21_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        psi_22_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        psi_23_curr = abs(np.random.multivariate_normal(np.zeros([shape]), sigma_psi * np.eye(shape)))
        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        res_psi21.append(psi_21_curr)
        res_psi22.append(psi_22_curr)
        res_psi23.append(psi_23_curr)

        # 3.sigma initialization
        res_sigma.append((self.sigma_post(a_e, b_e, n, SSE_new)).rvs())

        # 4.beta initialization
        res_beta1.append(self.Beta1)
        res_beta2.append(self.Beta2)

        # 5.PDE error  initialization
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  self.Beta1)  # zeta1 initialization
        zeta_new2 = self.zeta_cal(psi_21_curr, psi_22_curr, psi_23_curr, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                  self.Beta2)  # zeta2 initialization
        zeta_new = zeta_new1.T @ zeta_new1 + zeta_new2.T @ zeta_new2  # PDE error ssq
        print('zeta origin', zeta_new)
        zeta_burn.append(zeta_new)

        # 6.gamma initialization
        res_gamma.append(self.gamma_post(a_0, b_0, K, zeta_new).rvs())
        # iteration
        for i in range(N_burn):
            # 7.update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))

            # 8.update gamma
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma[-1])

            if i <= 100:
                F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1, res_psi11[-1], res_psi12[-1],
                                    res_psi13[-1])
                F_new2 = self.F_cal(self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2, res_psi21[-1], res_psi22[-1],
                                    res_psi23[-1])
                D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
                D_new2 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new2)
                # 9.update beta
                res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
                res_beta2.append(self.Beta_sample(res_beta2, D_new2, self.B_mat, self.Y2_obs, res_sigma[-1]))
            else:
                res_beta1.append(res_beta1[-1])
                res_beta2.append(res_beta2[-1])

            # 10.update time-varying parametersψ
            sample11, sample21 = self.psi_1(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            res_psi21.append(sample21)
            sample12, sample22 = self.psi_2(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            res_psi22.append(sample22)
            sample13, sample23 = self.psi_3(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)
            res_psi23.append(sample23)

            # 11.update Bspline error
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_2 = self.SSE_condition(self.Y2_obs, self.B_mat, res_beta2[-1])
            SSE_new = SSE_1 + SSE_2

            # 12.update PDE error
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], self.T_f_1, self.X_d_1, self.Y_d_1,
                                   self.Z_d_1, res_beta1[-1])
            zeta_2 = self.zeta_cal(res_psi21[-1], res_psi22[-1], res_psi23[-1], self.T_f_2, self.X_d_2, self.Y_d_2,
                                   self.Z_d_2, res_beta2[-1])
            zeta_new = zeta_1.T @ zeta_1 + zeta_2.T @ zeta_2
            zeta_burn.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new)
            print('------------------')

        self.sigma_e_2_burn = res_sigma[-1]
        self.gamma_0_burn = res_gamma[-1]
        self.beta1_burn = res_beta1[-1]
        self.beta2_burn = res_beta2[-1]
        self.psi_11_burn = res_psi11[-1]
        self.psi_12_burn = res_psi12[-1]
        self.psi_13_burn = res_psi13[-1]
        self.psi_21_burn = res_psi21[-1]
        self.psi_22_burn = res_psi22[-1]
        self.psi_23_burn = res_psi23[-1]

        print("sigma_e_2 estimation:", res_sigma[-1])
        print("gamma_0 estimation:", res_gamma[-1])
        print("beta1 estimation:", res_beta1[-1])
        print("beta2 estimation:", res_beta2[-1])

        print("psi_11 estimation:", res_psi11[-1])
        print("psi_12 estimation:", res_psi12[-1])
        print("psi_13 estimation:", res_psi13[-1])

        print("psi_21 estimation:", res_psi21[-1])
        print("psi_22 estimation:", res_psi22[-1])
        print("psi_23 estimation:", res_psi23[-1])
        return res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_beta1, res_beta2, res_sigma, res_gamma, zeta_burn

    def gibbs_normal(self, N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt_0):
        res_sigma = []
        res_gamma = []
        res_psi11 = []  # PDE1 2nd derivative
        res_psi12 = []  # PDE1 1st derivative
        res_psi13 = []  # PDE1 constant
        res_psi21 = []  # PDE2 2nd derivative
        res_psi22 = []  # PDE2 1st derivative
        res_psi23 = []  # PDE2 constant
        res_beta1 = []
        res_beta2 = []
        self.nt = nt_0
        zeta_normal = []  # zeta²

        # 1.Bspline error  initialization
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.beta1_burn)
        SSE_new2 = self.SSE_condition(self.Y2_obs, self.B_mat, self.beta2_burn)
        SSE_new = SSE_new1 + SSE_new2

        # 2.time-varying parameters initialization
        psi_11_curr = self.psi_11_burn
        psi_12_curr = self.psi_12_burn
        psi_13_curr = self.psi_13_burn

        psi_21_curr = self.psi_21_burn
        psi_22_curr = self.psi_22_burn
        psi_23_curr = self.psi_23_burn

        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        res_psi21.append(psi_21_curr)
        res_psi22.append(psi_22_curr)
        res_psi23.append(psi_23_curr)

        # 3.sigma initialization
        res_sigma.append(self.sigma_e_2_burn)

        # 4.beta  initialization
        res_beta1.append(self.beta1_burn)
        res_beta2.append(self.beta2_burn)

        # 5.PDE error  initialization
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1,
                                  res_beta1[-1])  # zeta initialization
        zeta_new2 = self.zeta_cal(psi_21_curr, psi_22_curr, psi_23_curr, self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2,
                                  res_beta2[-1])  # zeta2 initialization
        zeta_new = zeta_new1.T @ zeta_new1 + zeta_new2.T @ zeta_new2
        zeta_normal.append(zeta_new)

        # 6.gamma  initialization
        res_gamma.append(self.gamma_0_burn)

        for i in range(N_normal):
            # 7.update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))

            # 8.update gamma
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma[-1])

            # F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.Y_d_1, self.Z_d_1, res_psi11[-1], res_psi12[-1], res_psi13[-1])
            # F_new2 = self.F_cal(self.T_f_2, self.X_d_2, self.Y_d_2, self.Z_d_2, res_psi21[-1], res_psi22[-1], res_psi23[-1])
            # D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
            # D_new2 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new2)
            # 9. update beta
            # res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            # res_beta2.append(self.Beta_sample(res_beta2, D_new2, self.B_mat, self.Y2_obs, res_sigma[-1]))
            res_beta1.append(self.beta1_burn)
            res_beta2.append(self.beta2_burn)

            # 10.update time-varying parametersψ
            sample11, sample21 = self.psi_1(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            res_psi21.append(sample21)

            sample12, sample22 = self.psi_2(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            res_psi22.append(sample22)

            sample13, sample23 = self.psi_3(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1],
                                            res_psi23[-1], res_beta1[-1], res_beta2[-1], sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)
            res_psi23.append(sample23)

            # 11.update Bspline error
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_2 = self.SSE_condition(self.Y2_obs, self.B_mat, res_beta2[-1])
            SSE_new = SSE_1 + SSE_2

            # 11.update PDE error
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], self.T_f_1, self.X_d_1, self.Y_d_1,
                                   self.Z_d_1, res_beta1[-1])
            zeta_2 = self.zeta_cal(res_psi21[-1], res_psi22[-1], res_psi23[-1], self.T_f_2, self.X_d_2, self.Y_d_2,
                                   self.Z_d_2, res_beta2[-1])
            zeta_new = zeta_1.T @ zeta_1 + zeta_2.T @ zeta_2
            zeta_normal.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new)
            print('------------------')

        self.sigma_e_2_normal = res_sigma[-1]
        self.gamma_0_normal = res_gamma[-1]
        self.beta1_normal = res_beta1[-1]
        self.beta2_normal = res_beta2[-1]
        self.psi_11_normal = res_psi11[-1]
        self.psi_12_normal = res_psi12[-1]
        self.psi_13_normal = res_psi13[-1]
        self.psi_21_normal = res_psi21[-1]
        self.psi_22_normal = res_psi22[-1]
        self.psi_23_normal = res_psi23[-1]

        print("sigma_e_2 estimation:", res_sigma[-1])
        print("gamma_0 estimation:", res_gamma[-1])
        print("beta1 estimation:", res_beta1[-1])
        print("beta2 estimation:", res_beta2[-1])

        print("psi_11 estimation:", res_psi11[-1])
        print("psi_12 estimation:", res_psi12[-1])
        print("psi_13 estimation:", res_psi13[-1])

        print("psi_21 estimation:", res_psi21[-1])
        print("psi_22 estimation:", res_psi22[-1])
        print("psi_23 estimation:", res_psi23[-1])
        return res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_beta1, res_beta2, res_sigma, res_gamma, zeta_normal


# %%
# True value
B_psi = np.load('input/B_psi_91.npy')

Y_t = np.load('input/Y_temp.npy')
B_t = np.load('input/B_mat_temp.npy')

beta_ini_t = np.load('input/beta_temp.npy')
Y_hat_t = B_t @ beta_ini_t

X_d_t = np.load('input/X_d_temp.npy')
Y_d_t = np.load('input/Y_d_temp.npy')
Z_d_t = np.load('input/Z_d_temp.npy')
T_f_t = np.load('input/T_f_temp.npy')
T_f_t = T_f_t

Y_h = np.load('input/Y_hum.npy')
B_h = np.load('input/B_mat_hum.npy')  # B_h = B_t

beta_ini_h = np.load('input/beta_hum.npy')
Y_hat_h = B_h @ beta_ini_h

X_d_h = np.load('input/X_d_hum.npy')
Y_d_h = np.load('input/Y_d_hum.npy')
Z_d_h = np.load('input/Z_d_hum.npy')
T_f_h = np.load('input/T_f_hum.npy')
T_f_h = T_f_h

n = B_h.shape[0]  # observation
K = beta_ini_h.shape[0]  # num of basis funtions
a_e = 0.001
b_e = 0.001
a_0 = 0.001
b_0 = 0.001
sigma_psi = 1

sigma_prop1 = 0.01
sigma_prop2 = 0.0001

nt = 54
shape_psi = B_psi.shape[1]
N_burn = 2000
N_normal = 10000
# %%
start_time11 = time.time()
GibbsObj_1 = GibbsSampler(Y_t, Y_h, B_t, beta_ini_t, beta_ini_h, X_d_t, Y_d_t, Z_d_t, T_f_t, X_d_h, Y_d_h, Z_d_h, T_f_h,
                          B_psi)
res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_beta1, res_beta2, res_sigma1, res_gamma1, zeta_burn = GibbsObj_1.gibbs_burn(
    N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop1, nt, shape_psi)
end_time11 = time.time()
print("PDE 1&2 burn_in stage time:", end_time11 - start_time11, "s")
# %%
# PDE1 steady stage
start_time12 = time.time()
res_psi11_normal, res_psi12_normal, res_psi13_normal, res_psi21_normal, res_psi22_normal, res_psi23_normal, res_beta1_normal, res_beta2_normal, res_sigma1_normal, res_gamma1_normal, zeta_normal = GibbsObj_1.gibbs_normal(
    N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop2, nt)
end_time12 = time.time()
print("PDE 1&2 steady_stage time:", end_time12 - start_time12, "s")
# %%
np.save('output_2in1/psi_11_burn.npy', res_psi11)
np.save('output_2in1/psi_12_burn.npy', res_psi12)
np.save('output_2in1/psi_13_burn.npy', res_psi13)
np.save('output_2in1/beta1_burn.npy', res_beta1)
np.save('output_2in1/psi_21_burn.npy', res_psi21)
np.save('output_2in1/psi_22_burn.npy', res_psi22)
np.save('output_2in1/psi_23_burn.npy', res_psi23)
np.save('output_2in1/beta2_burn.npy', res_beta2)

np.save('output_2in1/sigma1_burn.npy', res_sigma1)
np.save('output_2in1/gamma1_burn.npy', res_gamma1)

np.save('output_2in1/psi_11_steady.npy', res_psi11_normal)
np.save('output_2in1/psi_12_steady.npy', res_psi12_normal)
np.save('output_2in1/psi_13_steady.npy', res_psi13_normal)
np.save('output_2in1/beta1_steady.npy', res_beta1_normal)
np.save('output_2in1/psi_21_steady.npy', res_psi21_normal)
np.save('output_2in1/psi_22_steady.npy', res_psi22_normal)
np.save('output_2in1/psi_23_steady.npy', res_psi23_normal)
np.save('output_2in1/beta2_steady.npy', res_beta2_normal)

np.save('output_2in1/sigma1_steady.npy', res_sigma1_normal)
np.save('output_2in1/gamma1_steady.npy', res_gamma1_normal)
np.save('output_2in1/zeta_burn.npy', zeta_burn)
np.save('output_2in1/zeta_steady.npy', zeta_normal)
# %%
np.save('output_2in1/sigma_sample.npy', GibbsObj_1.sigma_set)
np.save('output_2in1/gamma_sample.npy', GibbsObj_1.gamma_set)
np.save('output_2in1/psi11_sample.npy', GibbsObj_1.psi11_set)
np.save('output_2in1/psi12_sample.npy', GibbsObj_1.psi12_set)
np.save('output_2in1/psi13_sample.npy', GibbsObj_1.psi13_set)

np.save('output_2in1/psi21_sample.npy', GibbsObj_1.psi21_set)
np.save('output_2in1/psi22_sample.npy', GibbsObj_1.psi22_set)
np.save('output_2in1/psi23_sample.npy', GibbsObj_1.psi23_set)
# %%
psi_11_hat = res_psi11_normal[-1]
psi_12_hat = res_psi12_normal[-1]
psi_13_hat = res_psi13_normal[-1]
psi_21_hat = res_psi21_normal[-1]
psi_22_hat = res_psi22_normal[-1]
psi_23_hat = res_psi23_normal[-1]

beta1_hat = res_beta1_normal[-1]
beta2_hat = res_beta2_normal[-1]
sigma_hat = res_sigma1_normal[-1]
gamma_hat = res_gamma1_normal[-1]
# %%
# PDE error
zeta_ob1 = GibbsObj_1.zeta_cal(psi_11_hat, psi_12_hat, psi_13_hat, T_f_t, X_d_t, Y_d_t, Z_d_t, beta1_hat)
zeta_ob2 = GibbsObj_1.zeta_cal(psi_21_hat, psi_22_hat, psi_23_hat, T_f_h, X_d_h, Y_d_h, Z_d_h, beta2_hat)

zeta_ob_ss = zeta_ob1.T @ zeta_ob1 + zeta_ob2.T @ zeta_ob2

print('mean_zeta', 1 / 2 * (np.mean(zeta_ob1) + np.mean(zeta_ob2)))
print('mean_abs_zeta', 1 / 2 * (np.mean(np.abs(zeta_ob1)) + np.mean(np.abs(zeta_ob2))))
print('max_zeta', np.max([np.max(zeta_ob1), np.max(zeta_ob2)]))
print('min_zeta', np.min([np.min(zeta_ob1), np.min(zeta_ob2)]))
print('rmse_zeta', np.sqrt(zeta_ob_ss / zeta_ob1.shape[0] / 2))
print('SST', zeta_ob_ss)
# %%
# 统Bspline error
error_bs1 = Y_t - B_t @ beta1_hat
error_bs2 = Y_h - B_h @ beta2_hat
print(error_bs1.T @ error_bs1, error_bs2.T @ error_bs2)
error_ss = error_bs1.T @ error_bs1 + error_bs2.T @ error_bs2

print('mean_BSpline', 1 / 2 * (np.mean(error_bs1) + np.mean(error_bs2)))
print('mean_abs_BSpline', 1 / 2 * (np.mean(np.abs(error_bs1)) + np.mean(np.abs(error_bs2))))
print('max_BSpline', np.max([np.max(error_bs1), np.max(error_bs2)]))
print('min_BSpline', np.min([np.min(error_bs1), np.min(error_bs2)]))
print('rmse_BSpline', np.sqrt(error_ss / error_bs1.shape[0] / 2))
print('SST_BSpline', error_ss)
# %%
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import datetime

now = datetime.datetime.now()

wb = Workbook()
dest_filename = "%s_Temperature&Humidity_Time_varying_2in1_result.xlsx" % (now.strftime('%Y%m%d-%H%M%S'))

ws = wb.active
ws.title = 'Result'

ws['A1'] = "Date"
ws['B1'] = now.strftime('%Y-%m-%d')  # %H:%M:%S
ws['C1'] = "Result"
ws.merge_cells(range_string='C1:E1')

ws.append(["psi11", '采样估计值'] + [psi_11_hat[i] for i in range(shape_psi)])
ws.append(["psi12", '采样估计值'] + [psi_12_hat[i] for i in range(shape_psi)])
ws.append(["psi13", '采样估计值'] + [psi_13_hat[i] for i in range(shape_psi)])

ws.append(["psi21", '采样估计值'] + [psi_21_hat[i] for i in range(shape_psi)])
ws.append(["psi22", '采样估计值'] + [psi_22_hat[i] for i in range(shape_psi)])
ws.append(["psi23", '采样估计值'] + [psi_23_hat[i] for i in range(shape_psi)])

ws.append(["sigma_2", '采样估计值', sigma_hat, (b_e + error_ss / 4) / (a_e + n / 2 - 1)])
ws.append(["gamma", '采样估计值', gamma_hat, (a_0 + K / 2) / (b_0 + zeta_ob_ss / 4)])

ws['I1'] = "Model Errors"
ws.merge_cells(range_string='I1:L1')

ws['I2'] = "PDE temperature"
ws.merge_cells(range_string='I2:I7')
ws['I14'] = "PDE humidity"
ws.merge_cells(range_string='I14:I19')

ws['I8'] = "统计模型 temperature"
ws.merge_cells(range_string='I8:I13')
ws['I20'] = "统计模型 humidity"
ws.merge_cells(range_string='I20:I25')

J = ['mean_zeta', 'mean_abs_zeta', 'max_zeta', 'min_zeta', 'rmse_zeta', 'SST_zeta', 'mean_BSpline', 'mean_abs_BSpline',
     'max_BSpline', 'min_BSpline', 'rmse_BSpline', 'SST_BSpline']
K1 = ['均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和', '均值', '绝对值均值', '最大值', '最小值',
      '均方根误差', '总误差平方和']

L1 = [np.mean(zeta_ob1), np.mean(np.abs(zeta_ob1)), np.max(zeta_ob1), np.min(zeta_ob1),
      np.sqrt(zeta_ob1.T @ zeta_ob1 / zeta_ob1.shape[0]), zeta_ob1.T @ zeta_ob1, np.mean(error_bs1),
      np.mean(np.abs(error_bs1)), np.max(error_bs1), np.min(error_bs1),
      np.sqrt(error_bs1.T @ error_bs1 / error_bs1.shape[0]), error_bs1.T @ error_bs1]
L2 = [np.mean(zeta_ob2), np.mean(np.abs(zeta_ob2)), np.max(zeta_ob2), np.min(zeta_ob2),
      np.sqrt(zeta_ob2.T @ zeta_ob2 / zeta_ob2.shape[0]), zeta_ob2.T @ zeta_ob2, np.mean(error_bs2),
      np.mean(np.abs(error_bs2)), np.max(error_bs2), np.min(error_bs2),
      np.sqrt(error_bs2.T @ error_bs2 / error_bs2.shape[0]), error_bs2.T @ error_bs2]

for i in range(2, 14):
    ws['J%d' % i] = J[i - 2]
    ws['K%d' % i] = K1[i - 2]
    ws['L%d' % i] = L1[i - 2]

    ws['J%d' % (i + 12)] = J[i - 2]
    ws['K%d' % (i + 12)] = K1[i - 2]
    ws['L%d' % (i + 12)] = L2[i - 2]

ws['K26'] = "PDE_SST"
ws['K27'] = "B_Spline_SST"

ws['L26'] = zeta_ob1.T @ zeta_ob1 + zeta_ob2.T @ zeta_ob2
ws['L27'] = error_bs1.T @ error_bs1 + error_bs2.T @ error_bs2

N = ['#', 'Ver.', 'p_Bspline', 'p_Bpsi', 'n', 'K', 'a_e', 'b_e', 'a_0', 'b_0', 'sigma_psi', 'sigma_prop1',
     'sigma_prop2', 'nt', 'N_burn', 'N_normal']
O = ['Value', 'Granary_2in1', '3', '3', n, K, a_e, b_e, a_0, b_0, sigma_psi, sigma_prop1, sigma_prop2, nt, N_burn,
     N_normal]
P = ['含义', '版本', '统计建模样条阶数（次数+1）', '时变参数样条阶数（次数+1）', '观测值个数', '基函数个数', '', '', '', '',
     '', '第一阶段采样sigma（Burn-in stage）', '第二阶段采样sigma（Steady stage）', '', 'Burn-in 阶段采样次数',
     'Steady阶段采样次数']

for i in range(1, 17):
    ws['N%d' % i] = N[i - 1]
    ws['O%d' % i] = O[i - 1]
    ws['P%d' % i] = P[i - 1]


def setCellStyle(st):
    """
    设置单元格格式
    params: st:就是第一步创建的工作表
    """

    # 边框
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'),
    )

    # 对齐
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        indent=0
    )

    # 字体
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        strike=False,
        color='000000'
    )

    for row, row_ind in zip(st.iter_rows(), range(1, st.max_row + 1)):
        for cell in row:
            # 设置边框
            st[cell.coordinate].border = border
            # 设置居中对齐
            st[cell.coordinate].alignment = alignment
            # 行高40
            st.row_dimensions[row_ind].height = 14
            # 设置字体
            st[cell.coordinate].font = font

    # 设置列宽
    st.column_dimensions['A'].width = 6
    st.column_dimensions['B'].width = 14
    st.column_dimensions['C'].width = 14
    st.column_dimensions['D'].width = 14
    st.column_dimensions['E'].width = 14
    st.column_dimensions['F'].width = 14
    st.column_dimensions['G'].width = 14

    st.column_dimensions['I'].width = 14
    st.column_dimensions['J'].width = 14
    st.column_dimensions['k'].width = 14
    st.column_dimensions['L'].width = 14
    st.column_dimensions['N'].width = 14
    st.column_dimensions['O'].width = 35
    st.column_dimensions['P'].width = 35

    st.column_dimensions['H'].width = 2
    st.column_dimensions['M'].width = 2


setCellStyle(ws)
wb.save(dest_filename)
wb.close()