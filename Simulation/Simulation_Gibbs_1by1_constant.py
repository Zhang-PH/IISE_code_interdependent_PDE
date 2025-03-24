import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import invgamma, gamma, norm, multivariate_normal
import copy
import time
import openpyxl

class GibbsSampler:
    def __init__(self, Y1_obs, B_mat, Beta1, X_f_1, X_d_1, T_f_1):
        self.Y1_obs= Y1_obs
        self.B_mat= B_mat
        self.Y1_hat = Y1_obs@B_mat
        self.Beta1 = Beta1
        self.X_f_1 = X_f_1
        self.X_d_1 = X_d_1
        self.T_f_1 = T_f_1
        # self.B_psi = B_psi
        
        self.psi_11_burn = None
        self.psi_12_burn = None
        self.psi_13_burn = None
        
        self.psi_11_normal = None
        self.psi_12_normal = None
        self.psi_13_normal = None

        self.beta1_burn = None
        self.beta1_normal = None
        self.gamma_0_burn = None
        self.sigma_e_2_burn = None
        self.gamma_0_normal = None
        self.sigma_e_2_normal = None
        self.nt = None

    @staticmethod
    def SSE_condition(Y_curr, B_curr, beta_curr):
        return (Y_curr - B_curr@beta_curr).T@(Y_curr - B_curr@beta_curr)

    @staticmethod
    def gamma_post(a_0, b_0, n, zeta_curr):#gamma posterior
        return gamma(a = a_0 + n/2, scale = 1/(b_0 + zeta_curr/2))

    @staticmethod
    def sigma_post(a_e, b_e, n, SSE_curr):#sigma posterior
        return invgamma(a=a_e + n/2, scale=b_e + SSE_curr/2)
    
    def sigma_2(self, res_sigma, a_e, b_e, n, SSE_new):#sampling sigma_2
        states = []
        sigma_dist = self.sigma_post(a_e, b_e, n, SSE_new)
        cur = res_sigma[-1] # current
              
        next1 = sigma_dist.rvs() #new
        exp1 = sigma_dist.pdf(next1)
        exp2 = sigma_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]    

    def gamma_0(self, res_gamma, a_0, b_0, K, zeta_curr):#sampling gamma_0
        states = []
        gamma_dist = self.gamma_post(a_0, b_0, K, zeta_curr)
        cur =  res_gamma[-1] # current
        next1 = gamma_dist.rvs() #new
        exp1 = gamma_dist.pdf(next1)
        exp2 = gamma_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]     

    def F_cal(self, T_f, X_d, X_f, B, psi_1_curr, psi_2_curr, psi_3_curr):
        return (T_f - psi_1_curr*X_d - psi_2_curr*X_f - psi_3_curr*B)

    @staticmethod
    def D_cal(B, sigma_curr, gamma_curr, F_curr):
        return (np.linalg.inv(B.T@B + sigma_curr*gamma_curr*F_curr.T@F_curr))

    @staticmethod
    def beta_post(D, B, Y, sigma):#sigma posterior
        return multivariate_normal(mean=D@B.T@Y, cov=sigma*D)

    def Beta_sample(self, res_beta, D, B, Y, sigma):#sampling beta
        states = []
        beta_dist = self.beta_post(D, B, Y, sigma)

        cur =  res_beta[-1] # current
        for i in range(1):
            next1 = beta_dist.rvs() #new

            exp1 = beta_dist.pdf(next1)
            exp2 = beta_dist.pdf(cur)
            if exp2 < exp1:
                states.append(next1)
            else:
                states.append(cur)
        return states[-1]
    
    @staticmethod
    def psi_single(psi_i, sigma_prop):#psi posterior
        return norm(loc=psi_i, scale=sigma_prop)
    
    def zeta_cal(self, tm1, tm2, tm3, T_f, X_d, X_f, beta):
        zeta = (T_f - tm1*X_d - tm2*X_f - tm3*self.B_mat)@beta
        return zeta
    
    def psi_1(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):#sampling psi_x1
        states1 = []

        t11_cur = res_t11[-1] # current
        print('psi_m1:')
        print(t11_cur)
        t12_cur = res_t12[-1]
        t13_cur = res_t13[-1]       
        
        new1 = (self.psi_single(t11_cur, sigma_prop).rvs()) #new
        next_t11 = new1
        
        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1[-1])
        zeta_new1 = self.zeta_cal(next_t11, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1[-1])

        zeta_sumsq_old = zeta_old1.T@zeta_old1 #+ zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
        zeta_sumsq_new = zeta_new1.T@zeta_new1 #+ zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
        
        square_cur = (t11_cur**2 + t12_cur**2 + t13_cur**2)/(2*sigma_psi)
        square_new = (next_t11**2 + t12_cur**2 + t13_cur**2)/(2*sigma_psi)
        
        exp1 = (-square_cur/2 - gamma_0*zeta_sumsq_old/2)
        exp2 = (-square_new/2 - gamma_0*zeta_sumsq_new/2)
        if exp2 > exp1:
            states1.append(next_t11)
            t11_cur = next_t11
        else:
            states1.append(t11_cur)
            
        print(zeta_sumsq_old)
        return states1[-1]
    
    def psi_2(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):#sampling psi_x2
        states1 = []

        print('psi_m2:')
        t11_cur = res_t11[-1]
        t12_cur = res_t12[-1] # current
        t13_cur = res_t13[-1]
        print(t12_cur)
        
        new1 = (self.psi_single(t12_cur, sigma_prop).rvs()) #new
        next_t12 = new1
        
        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1[-1])
        zeta_new1 = self.zeta_cal(t11_cur, next_t12, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1[-1])

        zeta_sumsq_old = zeta_old1.T@zeta_old1
        zeta_sumsq_new = zeta_new1.T@zeta_new1
        
        square_cur = (t11_cur**2 + t12_cur**2 + t13_cur**2)/(2*sigma_psi)
        square_new = (t11_cur**2 + next_t12**2 + t13_cur**2)/(2*sigma_psi)
        
        exp1 = (-square_cur/2 - gamma_0*zeta_sumsq_old/2)
        exp2 = (-square_new/2 - gamma_0*zeta_sumsq_new/2)
        if exp2 > exp1:
            states1.append(next_t12)
            t12_cur = next_t12
        else:
            states1.append(t12_cur)
            
        print(zeta_sumsq_old)
        return states1[-1]
    
    def psi_3(self, res_t11, res_t12, res_t13, res_beta1, sigma_prop, gamma_0):#sampling psi_x3
        states1 = []

        print('psi_m3:')
        t11_cur = res_t11[-1]
        t12_cur = res_t12[-1] 
        t13_cur = res_t13[-1]# current
        print(t13_cur)

        new1 = (self.psi_single(t13_cur, sigma_prop).rvs()) #new
        next_t13 = new1
        
        zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1[-1])
        zeta_new1 = self.zeta_cal(t11_cur, t12_cur, next_t13, T_f_1, X_d_1, X_f_1, res_beta1[-1])
        
        zeta_sumsq_old = zeta_old1.T@zeta_old1
        zeta_sumsq_new = zeta_new1.T@zeta_new1
        
        square_cur = (t11_cur**2 + t12_cur**2 + t13_cur**2)/(2*sigma_psi)
        square_new = (t11_cur**2 + t12_cur**2 + next_t13**2)/(2*sigma_psi)
        
        exp1 = (-square_cur/2 - gamma_0*zeta_sumsq_old/2)
        exp2 = (-square_new/2 - gamma_0*zeta_sumsq_new/2)
        if exp2 > exp1:
            states1.append(next_t13)
            # t13_cur = next_t13
        else:
            states1.append(t13_cur)
        print(zeta_sumsq_old)
        return states1[-1]
    
    def gibbs_burn(self, N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt, shape):
        res_sigma = []
        res_gamma = []
        res_psi11 = []#PDE1 2nd derivative
        res_psi12 = []#PDE1 1st derivative
        res_psi13 = []#PDE1 constant
        res_beta1 = []
        self.nt = nt
        
        #initialization
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.Beta1)
        SSE_new = SSE_new1
        
        psi_11_curr = np.random.normal(0,sigma_psi)
        psi_12_curr = np.random.normal(0,sigma_psi)
        psi_13_curr = np.random.normal(0,sigma_psi)

        res_sigma.append((self.sigma_post(a_e, b_e, n, SSE_new)).rvs())#sigma
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, T_f_1, X_d_1, X_f_1, self.Beta1) #zeta

        zeta_new = zeta_new1.T@zeta_new1# + zeta_new2.T@zeta_new2 + zeta_new3.T@zeta_new3
        
        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        
        res_gamma.append(self.gamma_post(a_0, b_0, K, zeta_new).rvs())#gamma
        res_beta1.append(self.Beta1)

        for i in range(N_burn):
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.X_f_1, self.B_mat, res_psi11[-1], res_psi12[-1], res_psi13[-1])
            D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
            res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))

            sample11 = self.psi_1(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            
            sample12 = self.psi_2(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            
            sample13 = self.psi_3(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)

            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_new = SSE_1
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], T_f_1, X_d_1, X_f_1, res_beta1[-1])
            zeta_new = zeta_1.T@zeta_1
            print(i)
        
        self.sigma_e_2_burn = res_sigma[-1]
        self.gamma_0_burn = res_gamma[-1]
        self.beta1_burn = res_beta1[-1]
        self.psi_11_burn = res_psi11[-1]
        self.psi_12_burn = res_psi12[-1]
        self.psi_13_burn = res_psi13[-1]

        print("sigma_e_2 estimation",res_sigma[-1])
        print("gamma_0 estimation",res_gamma[-1])
        print("beta1 estimation",res_beta1[-1])
        
        print("psi_1 estimation",res_psi11[-1])
        print("psi_2 estimation",res_psi12[-1])
        print("psi_3 estimation",res_psi13[-1])
        return res_psi11, res_psi12, res_psi13, res_beta1, res_sigma, res_gamma
    
    def gibbs_normal(self, N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop, nt_0):
        res_sigma = []
        res_gamma = []
        
        res_psi11 = []#PDE1 2nd derivative
        res_psi12 = []#PDE1 1st derivative
        res_psi13 = []#PDE1 constant
        
        res_beta1 = []

        self.nt = nt_0
        
        #initialization
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.beta1_burn)
        SSE_new = SSE_new1
        
        psi_11_curr = self.psi_11_burn
        psi_12_curr = self.psi_12_burn
        psi_13_curr = self.psi_13_burn

        res_sigma.append((self.sigma_post(a_e, b_e, n, SSE_new)).rvs())#sigma
        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        
        res_gamma.append(self.gamma_0_burn)#gamma
        res_beta1.append(self.beta1_burn)
        
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, T_f_1, X_d_1, X_f_1, res_beta1[-1]) #zeta

        zeta_new = zeta_new1.T@zeta_new1   

        for i in range(N_normal):
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))
            res_gamma.append(self.gamma_0(res_gamma, a_0, b_0, K, zeta_new))
            F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.X_f_1, self.B_mat, res_psi11[-1], res_psi12[-1], res_psi13[-1])
            D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma[-1], F_new1)
            res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            
            sample11 = self.psi_1(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi11.append(sample11)
            
            sample12 = self.psi_2(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi12.append(sample12)
            
            sample13 = self.psi_3(res_psi11, res_psi12, res_psi13, res_beta1, sigma_prop, res_gamma[-1])
            res_psi13.append(sample13)

            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_new = SSE_1
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], T_f_1, X_d_1, X_f_1, res_beta1[-1])
            zeta_new = zeta_1.T@zeta_1
            print(i)
        
        self.sigma_e_2_normal = res_sigma[-1]
        self.gamma_0_normal = res_gamma[-1]
        self.beta1_normal = res_beta1[-1]
        self.psi_11_normal = res_psi11[-1]
        self.psi_12_normal = res_psi12[-1]
        self.psi_13_normal = res_psi13[-1]

        print("sigma_e_2 estimation",res_sigma[-1])
        print("gamma_0 estimation",res_gamma[-1])
        print("beta1 estimation",res_beta1[-1])
        
        print("psi_1 estimation",res_psi11[-1])
        print("psi_2 estimation",res_psi12[-1])
        print("psi_3 estimation",res_psi13[-1])
        return res_psi11, res_psi12, res_psi13, res_beta1, res_sigma, res_gamma

#true value
Y1 = np.load('input_simulation/Y_TV_1.npy')#_small
B1 = np.load('input_simulation/B_mat_TV_1.npy')

beta1_ini = np.load('input_simulation/beta_TV_1.npy')
Y1_hat = B1@beta1_ini
X_f_1 = np.load('input_simulation/X_f_TV_1.npy')
X_d_1 = np.load('input_simulation/X_d_TV_1.npy')
T_f_1 = np.load('input_simulation/T_f_TV_1.npy')

Y2 = np.load('input_simulation/Y_TV_2.npy')#_small
B2 = np.load('input_simulation/B_mat_TV_2.npy')

beta2_ini = np.load('input_simulation/beta_TV_2.npy')
Y2_hat = B2@beta2_ini
X_f_2 = np.load('input_simulation/X_f_TV_2.npy')
X_d_2 = np.load('input_simulation/X_d_TV_2.npy')
T_f_2 = np.load('input_simulation/T_f_TV_2.npy')

Y3 = np.load('input_simulation/Y_TV_3.npy')#_small
B3 = np.load('input_simulation/B_mat_TV_3.npy')

beta3_ini = np.load('input_simulation/beta_TV_3.npy')
Y3_hat = B3@beta3_ini
X_f_3 = np.load('input_simulation/X_f_TV_3.npy')
X_d_3 = np.load('input_simulation/X_d_TV_3.npy')
T_f_3 = np.load('input_simulation/T_f_TV_3.npy')

B_psi = np.load('input_simulation/B_psi.npy')
psi_node = np.load('input_simulation/psi_node.npy')

H11 = np.load('input_simulation/H11.npy')
H12 = np.load('input_simulation/H12.npy')

H21 = np.load('input_simulation/H21.npy')
H22 = np.load('input_simulation/H22.npy')

H31 = np.load('input_simulation/H31.npy')
H32 = np.load('input_simulation/H32.npy')

H1 = 1e-3*H11 + 1e-6*H12
H2 = 1e-3*H21 + 1e-6*H22
H3 = 1e-3*H31 + 1e-6*H32
n = B1.shape[0] #observation
K = beta1_ini.shape[0] #basis function

order = 4
order_penalty = 3
lambda_penalty = 1e-3
sigma_noise = 0

a_e = 1.01
b_e = 1e-6

a_0 = 1e-7
b_0 = 1e-7

a_1 = 1e-7
b_1 = 1e-7

a_2 = 1e-7
b_2 = 1e-7
sigma_psi = 3#1

sigma_prop1 = 0.01
sigma_prop2 = 0.01

nt = 33
shape_psi = B_psi.shape[1] 
N_burn = 5000
N_normal = 5000

#PDE1 burn_in stage
start_time11 = time.time()
GibbsObj_1 = GibbsSampler(Y1, B1, beta1_ini, X_f_1, X_d_1, T_f_1)
res_psi11, res_psi12, res_psi13, res_beta1, res_sigma1, res_gamma1 = GibbsObj_1.gibbs_burn(N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop1, nt, shape_psi)
end_time11 = time.time()
print("PDE1 burn_in stage time：", end_time11 - start_time11, "s")

#PDE1 steady stage
start_time12 = time.time()
res_psi11_normal, res_psi12_normal, res_psi13_normal, res_beta1_normal, res_sigma1_normal, res_gamma1_normal = GibbsObj_1.gibbs_normal(N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop2, nt)
end_time12 = time.time()
print("PDE1 steady_stage time：", end_time12 - start_time12, "s")

#PDE2 burn_in stage
start_time21 = time.time()
GibbsObj_2 = GibbsSampler(Y2, B2, beta2_ini, X_f_2, X_d_2, T_f_2)
res_psi21, res_psi22, res_psi23, res_beta2, res_sigma2, res_gamma2 = GibbsObj_2.gibbs_burn(N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop1, nt, shape_psi)
end_time21 = time.time()
print("PDE2 burn_in stage time：", end_time21 - start_time21, "s")

#PDE2 steady stage
start_time22 = time.time()
res_psi21_normal, res_psi22_normal, res_psi23_normal, res_beta2_normal, res_sigma2_normal, res_gamma2_normal = GibbsObj_2.gibbs_normal(N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop2, nt)
end_time22 = time.time()
print("PDE2 steady_stage time：", end_time22 - start_time22, "s")

#PDE3 burn_in stage
start_time31 = time.time()
GibbsObj_3 = GibbsSampler(Y3, B3, beta3_ini, X_f_3, X_d_3, T_f_3)
res_psi31, res_psi32, res_psi33, res_beta3, res_sigma3, res_gamma3 = GibbsObj_3.gibbs_burn(N_burn, n, K, a_e, b_e, a_0, b_0, sigma_prop1, nt, shape_psi)
end_time31 = time.time()
print("PDE3 burn_in stage time：", end_time31 - start_time31, "s")

#PDE3 steady stage
start_time32 = time.time()
res_psi31_normal, res_psi32_normal, res_psi33_normal, res_beta3_normal, res_sigma3_normal, res_gamma3_normal = GibbsObj_3.gibbs_normal(
    N_normal, n, K, a_e, b_e, a_0, b_0, sigma_prop2, nt)
end_time32 = time.time()
print("PDE3 steady_stage time：", end_time32 - start_time32, "s")

psi_11_hat = res_psi11_normal[-1]
psi_12_hat = res_psi12_normal[-1]
psi_13_hat = res_psi13_normal[-1]
psi_21_hat = res_psi21_normal[-1]
psi_22_hat = res_psi22_normal[-1]
psi_23_hat = res_psi23_normal[-1]
psi_31_hat = res_psi31_normal[-1]
psi_32_hat = res_psi32_normal[-1]
psi_33_hat = res_psi33_normal[-1]
beta1_hat = res_beta1_normal[-1]
beta2_hat = res_beta2_normal[-1]
beta3_hat = res_beta3_normal[-1]

sigma1_hat = res_sigma1_normal[-1]
gamma1_hat = res_gamma1_normal[-1]
sigma2_hat = res_sigma2_normal[-1]
gamma2_hat = res_gamma2_normal[-1]
sigma3_hat = res_sigma3_normal[-1]
gamma3_hat = res_gamma3_normal[-1]

#PDE error
zeta_ob1 = GibbsObj_1.zeta_cal(psi_11_hat, psi_12_hat, psi_13_hat, T_f_1, X_d_1, X_f_1, beta1_hat)
zeta_ob2 = GibbsObj_1.zeta_cal(psi_21_hat, psi_22_hat, psi_23_hat, T_f_2, X_d_2, X_f_2, beta2_hat)
zeta_ob3 = GibbsObj_1.zeta_cal(psi_31_hat, psi_32_hat, psi_33_hat, T_f_3, X_d_3, X_f_3, beta3_hat)
zeta_ob_ss = zeta_ob1.T@zeta_ob1 + zeta_ob2.T@zeta_ob2 + zeta_ob3.T@zeta_ob3

print('mean_zeta', 1/3*(np.mean(zeta_ob1) + np.mean(zeta_ob2) + np.mean(zeta_ob3)))
print('mean_abs_zeta', 1/3*(np.mean(np.abs(zeta_ob1)) + np.mean(np.abs(zeta_ob2)) + np.mean(np.abs(zeta_ob3))))
print('max_zeta', np.max([np.max(zeta_ob1), np.max(zeta_ob2), np.max(zeta_ob3)]))
print('min_zeta', np.min([np.min(zeta_ob1), np.min(zeta_ob2), np.min(zeta_ob3)]))
print('rmse_zeta', np.sqrt(zeta_ob_ss/zeta_ob1.shape[0]/3))
print('SST', zeta_ob_ss)

#Bspline error
error_bs1 = Y1-B1@beta1_hat
error_bs2 = Y2-B2@beta2_hat
error_bs3 = Y3-B3@beta3_hat
error_ss = error_bs1.T@error_bs1 + error_bs2.T@error_bs2 + error_bs3.T@error_bs3

print('mean_BSpline', 1/3*(np.mean(error_bs1) + np.mean(error_bs2) + np.mean(error_bs3)))
print('mean_abs_BSpline', 1/3*(np.mean(np.abs(error_bs1)) + np.mean(np.abs(error_bs2)) + np.mean(np.abs(error_bs3))))
print('max_BSpline', np.max([np.max(error_bs1), np.max(error_bs2), np.max(error_bs3)]))
print('min_BSpline', np.min([np.min(error_bs1), np.min(error_bs2), np.min(error_bs3)]))
print('rmse_BSpline', np.sqrt(error_ss/error_bs1.shape[0]/3))
print('SST_BSpline', error_ss)

np.save('output_constant_simulation/psi_11_burn.npy',res_psi11)
np.save('output_constant_simulation/psi_12_burn.npy',res_psi12)
np.save('output_constant_simulation/psi_13_burn.npy',res_psi13)
np.save('output_constant_simulation/beta1_burn.npy',res_beta1)
np.save('output_constant_simulation/psi_21_burn.npy',res_psi21)
np.save('output_constant_simulation/psi_22_burn.npy',res_psi22)
np.save('output_constant_simulation/psi_23_burn.npy',res_psi23)
np.save('output_constant_simulation/beta2_burn.npy',res_beta2)
np.save('output_constant_simulation/psi_31_burn.npy',res_psi31)
np.save('output_constant_simulation/psi_32_burn.npy',res_psi32)
np.save('output_constant_simulation/psi_33_burn.npy',res_psi33)
np.save('output_constant_simulation/beta3_burn.npy',res_beta3)

np.save('output_constant_simulation/sigma1_burn.npy',res_sigma1)
np.save('output_constant_simulation/gamma1_burn.npy',res_gamma1)

np.save('output_constant_simulation/sigma2_burn.npy',res_sigma2)
np.save('output_constant_simulation/gamma2_burn.npy',res_gamma2)

np.save('output_constant_simulation/sigma3_burn.npy',res_sigma3)
np.save('output_constant_simulation/gamma3_burn.npy',res_gamma3)


np.save('output_constant_simulation/psi_11_steady.npy',res_psi11_normal)
np.save('output_constant_simulation/psi_12_steady.npy',res_psi12_normal)
np.save('output_constant_simulation/psi_13_steady.npy',res_psi13_normal)
np.save('output_constant_simulation/beta1_steady.npy',res_beta1_normal)
np.save('output_constant_simulation/psi_21_steady.npy',res_psi21_normal)
np.save('output_constant_simulation/psi_22_steady.npy',res_psi22_normal)
np.save('output_constant_simulation/psi_23_steady.npy',res_psi23_normal)
np.save('output_constant_simulation/beta2_steady.npy',res_beta2_normal)
np.save('output_constant_simulation/psi_31_steady.npy',res_psi31_normal)
np.save('output_constant_simulation/psi_32_steady.npy',res_psi32_normal)
np.save('output_constant_simulation/psi_33_steady.npy',res_psi33_normal)
np.save('output_constant_simulation/beta3_steady.npy',res_beta3_normal)

np.save('output_constant_simulation/sigma1_steady.npy',res_sigma1_normal)
np.save('output_constant_simulation/gamma1_steady.npy',res_gamma1_normal)
np.save('output_constant_simulation/sigma2_steady.npy',res_sigma2_normal)
np.save('output_constant_simulation/gamma2_steady.npy',res_gamma2_normal)
np.save('output_constant_simulation/sigma3_steady.npy',res_sigma3_normal)
np.save('output_constant_simulation/gamma3_steady.npy',res_gamma3_normal)

psi11_node = psi_node[0][0]
psi12_node = psi_node[0][1]
psi13_node = psi_node[0][2]
psi21_node = psi_node[1][0]
psi22_node = psi_node[1][1]
psi23_node = psi_node[1][2]
psi31_node = psi_node[2][0]
psi32_node = psi_node[2][1]
psi33_node = psi_node[2][2]


from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment #PatternFill
import datetime
now = datetime.datetime.now()

wb = Workbook()
dest_filename = "%s_constant_simulation_1by1_result.xlsx"%(now.strftime('%Y%m%d-%H%M%S'))

ws = wb.active
ws.title = 'Result'

ws['A1'] = "Date"
ws['B1'] = now.strftime('%Y-%m-%d')#%H:%M:%S
ws['C1'] = "Result"
ws.merge_cells(range_string='C1:E1')

ws.append(["psi11", '真值']+[psi11_node[i] for i in range(shape_psi)])
ws.append(["psi11", '采样估计值']+[psi_11_hat for i in range(shape_psi)])
ws.append(["psi11", '误差百分比']+[abs((psi11_node[i]-psi_11_hat)/psi11_node[i]) for i in range(shape_psi)])

ws.append(["psi12", '真值']+[psi12_node[i] for i in range(shape_psi)])
ws.append(["psi12", '采样估计值']+[psi_12_hat for i in range(shape_psi)])
ws.append(["psi12", '误差百分比']+[abs((psi12_node[i]-psi_12_hat)/psi12_node[i]) for i in range(shape_psi)])

ws.append(["psi13", '真值']+[psi13_node[i] for i in range(shape_psi)])
ws.append(["psi13", '采样估计值']+[psi_13_hat for i in range(shape_psi)])
ws.append(["psi13", '误差百分比']+[abs((psi13_node[i]-psi_13_hat)/psi13_node[i]) for i in range(shape_psi)])

ws.append(["psi21", '真值']+[psi21_node[i] for i in range(shape_psi)])
ws.append(["psi21", '采样估计值']+[psi_21_hat for i in range(shape_psi)])
ws.append(["psi21", '误差百分比']+[abs((psi21_node[i]-psi_21_hat)/psi21_node[i]) for i in range(shape_psi)])

ws.append(["psi22", '真值']+[psi22_node[i] for i in range(shape_psi)])
ws.append(["psi22", '采样估计值']+[psi_22_hat for i in range(shape_psi)])
ws.append(["psi22", '误差百分比']+[abs((psi22_node[i]-psi_22_hat)/psi22_node[i]) for i in range(shape_psi)])

ws.append(["psi23", '真值']+[psi23_node[i] for i in range(shape_psi)])
ws.append(["psi23", '采样估计值']+[psi_23_hat for i in range(shape_psi)])
ws.append(["psi23", '误差百分比']+[abs((psi23_node[i]-psi_23_hat)/psi23_node[i]) for i in range(shape_psi)])

ws.append(["psi31", '真值']+[psi11_node[i] for i in range(shape_psi)])
ws.append(["psi31", '采样估计值']+[psi_31_hat for i in range(shape_psi)])
ws.append(["psi31", '误差百分比']+[abs((psi11_node[i]-psi_31_hat)/psi11_node[i]) for i in range(shape_psi)])

ws.append(["psi32", '真值']+[psi12_node[i] for i in range(shape_psi)])
ws.append(["psi32", '采样估计值']+[psi_32_hat for i in range(shape_psi)])
ws.append(["psi32", '误差百分比']+[abs((psi12_node[i]-psi_32_hat)/psi12_node[i]) for i in range(shape_psi)])

ws.append(["psi33", '真值']+[psi13_node[i] for i in range(shape_psi)])
ws.append(["psi33", '采样估计值']+[psi_33_hat for i in range(shape_psi)])
ws.append(["psi33", '误差百分比']+[abs((psi13_node[i]-psi_33_hat)/psi13_node[i]) for i in range(shape_psi)])

ws.append(["sigma1_2", '采样估计值', sigma1_hat, (b_e+error_bs1.T@error_bs1/2)/(a_e+n/2-1)])
ws.append(["gamma1", '采样估计值', gamma1_hat, (a_0+K/2)/(b_0+zeta_ob1.T@zeta_ob1/2)])
ws.append(["sigma2_2", '采样估计值', sigma2_hat, (b_e+error_bs2.T@error_bs2/2)/(a_e+n/2-1)])
ws.append(["gamma2", '采样估计值', gamma2_hat, (a_0+K/2)/(b_0+zeta_ob2.T@zeta_ob2/2)])
ws.append(["sigma3_2", '采样估计值', sigma3_hat, (b_e+error_bs3.T@error_bs3/2)/(a_e+n/2-1)])
ws.append(["gamma3", '采样估计值', gamma3_hat, (a_0+K/2)/(b_0+zeta_ob3.T@zeta_ob3/2)])

ws.merge_cells(range_string='A2:A4')
ws.merge_cells(range_string='A5:A7')
ws.merge_cells(range_string='A8:A10')

ws.merge_cells(range_string='A11:A13')
ws.merge_cells(range_string='A14:A16')
ws.merge_cells(range_string='A17:A19')

ws.merge_cells(range_string='A20:A22')
ws.merge_cells(range_string='A23:A25')
ws.merge_cells(range_string='A26:A28')

ws['I1'] = "Model Errors"
ws.merge_cells(range_string='I1:L1')

ws['I2'] = "PDE 1"
ws.merge_cells(range_string='I2:I7')
ws['I14'] = "PDE 2"
ws.merge_cells(range_string='I14:I19')
ws['I26'] = "PDE 3"
ws.merge_cells(range_string='I26:I31')

J = ['mean_zeta', 'mean_abs_zeta','max_zeta','min_zeta','rmse_zeta','SST_zeta', 'mean_BSpline','mean_abs_BSpline','max_BSpline','min_BSpline','rmse_BSpline','SST_BSpline']
K1 = ['均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和','均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和']

L1 = [np.mean(zeta_ob1), np.mean(np.abs(zeta_ob1)), np.max(zeta_ob1), np.min(zeta_ob1), np.sqrt(zeta_ob1.T@zeta_ob1/zeta_ob1.shape[0]), zeta_ob1.T@zeta_ob1,np.mean(error_bs1), np.mean(np.abs(error_bs1)), np.max(error_bs1), np.min(error_bs1), np.sqrt(error_bs1.T@error_bs1/error_bs1.shape[0]), error_bs1.T@error_bs1]

L2 = [np.mean(zeta_ob2), np.mean(np.abs(zeta_ob2)), np.max(zeta_ob2), np.min(zeta_ob2), np.sqrt(zeta_ob2.T@zeta_ob2/zeta_ob2.shape[0]), zeta_ob2.T@zeta_ob2,np.mean(error_bs2), np.mean(np.abs(error_bs2)), np.max(error_bs2), np.min(error_bs2), np.sqrt(error_bs2.T@error_bs2/error_bs2.shape[0]), error_bs2.T@error_bs2]

L3 = [np.mean(zeta_ob3), np.mean(np.abs(zeta_ob3)), np.max(zeta_ob3), np.min(zeta_ob3), np.sqrt(zeta_ob3.T@zeta_ob3/zeta_ob3.shape[0]), zeta_ob3.T@zeta_ob3,np.mean(error_bs3), np.mean(np.abs(error_bs3)), np.max(error_bs3), np.min(error_bs3), np.sqrt(error_bs3.T@error_bs3/error_bs3.shape[0]), error_bs3.T@error_bs3]

for i in range(2,14):
    ws['J%d'%i] = J[i-2]
    ws['K%d'%i] = K1[i-2]
    ws['L%d'%i] = L1[i-2]
    
    ws['J%d'%(i+12)] = J[i-2]
    ws['K%d'%(i+12)] = K1[i-2]
    ws['L%d'%(i+12)] = L2[i-2]
    
    ws['J%d'%(i+24)] = J[i-2]
    ws['K%d'%(i+24)] = K1[i-2]
    ws['L%d'%(i+24)] = L3[i-2]
    
ws['I8'] = "统计模型 1"
ws.merge_cells(range_string='I8:I13')
ws['I20'] = "统计模型 2"
ws.merge_cells(range_string='I20:I25')
ws['I32'] = "统计模型 3"
ws.merge_cells(range_string='I32:I37')

ws['K38'] = "PDE_SST"
ws['K39'] = "B_Spline_SST"

ws['L38'] = zeta_ob1.T@zeta_ob1 + zeta_ob2.T@zeta_ob2 + zeta_ob3.T@zeta_ob3
ws['L39'] = error_bs1.T@error_bs1 + error_bs2.T@error_bs2 + error_bs3.T@error_bs3

N = ['#', 'Ver.', 'p_Bspline', 'p_Bpsi', 'n', 'K', 'a_e', 'b_e', 'a_0', 'b_0', 'sigma_psi', 'sigma_prop1', 'sigma_prop2', 'nt', 'N_burn', 'N_normal']
O = ['Value', 'Constant 1by1', '3', '3', n, K, a_e, b_e, a_0, b_0, sigma_psi, sigma_prop1, sigma_prop2, nt, N_burn, N_normal]
P = ['含义', '版本', '统计建模样条阶数（次数+1）', '时变参数样条阶数（次数+1）', '观测值个数', '基函数个数', '', '', '', '', '', '第一阶段采样sigma（Burn-in stage）', '第二阶段采样sigma（Steady stage）', '', 'Burn-in 阶段采样次数', 'Steady阶段采样次数']

for i in range(1,17):
    ws['N%d'%i] = N[i-1]
    ws['O%d'%i] = O[i-1]
    ws['P%d'%i] = P[i-1]

def setCellStyle(st):
     """
     设置单元格格式
     params: st:就是第一步创建的工作表
     """

     # 边框
     border = Border(
         left=Side(border_style='thin', color='000000'),
         right=Side(border_style='thin', color='000000'),
         top=Side(border_style='thin', color='000000'),
         bottom=Side(border_style='thin', color='000000'),
     )

     # 对齐
     alignment = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0
     )

     # 字体     
     font = Font(
         name='Times New Roman',
         size=11,
         bold=False,
         italic=False,
         strike=False,
         color='000000'
     )

     for row, row_ind in zip(st.iter_rows(), range(1, st.max_row + 1)):
         for cell in row:
         	# 设置边框
            st[cell.coordinate].border = border
     		# 设置居中对齐
            st[cell.coordinate].alignment = alignment
            # 行高40
            st.row_dimensions[row_ind].height = 14
            # 设置字体
            st[cell.coordinate].font = font

      # 设置列宽         
     st.column_dimensions['A'].width = 6
     st.column_dimensions['B'].width = 14
     st.column_dimensions['C'].width = 14
     st.column_dimensions['D'].width = 14
     st.column_dimensions['E'].width = 14
     st.column_dimensions['F'].width = 14
     st.column_dimensions['G'].width = 14
  
     st.column_dimensions['I'].width = 14
     st.column_dimensions['J'].width = 14
     st.column_dimensions['k'].width = 14
     st.column_dimensions['L'].width = 14
     st.column_dimensions['N'].width = 14
     st.column_dimensions['O'].width = 35
     st.column_dimensions['P'].width = 35
     
     st.column_dimensions['H'].width = 2
     st.column_dimensions['M'].width = 2

setCellStyle(ws)
wb.save(dest_filename)
wb.close()