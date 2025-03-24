import numpy as np
from scipy.stats import invgamma, gamma, norm, multivariate_normal, pearsonr
import copy
from time import perf_counter

class GibbsSampler: #Ver. Simulation 3in1
    def __init__(self, Y1_obs, Y2_obs, Y3_obs, B_mat, Beta1, Beta2, Beta3, X_f_1, X_d_1, T_f_1, X_f_2, X_d_2, T_f_2, X_f_3, X_d_3, T_f_3, B_psi, H):
        self.Y1_obs= Y1_obs
        self.Y2_obs= Y2_obs
        self.Y3_obs= Y3_obs
        self.B_mat= B_mat
        self.Y1_hat = Y1_obs@B_mat
        self.Y2_hat = Y2_obs@B_mat
        self.Y3_hat = Y3_obs@B_mat
        self.Beta1 = Beta1
        self.Beta2 = Beta2
        self.Beta3 = Beta3
        self.X_f_1 = X_f_1
        self.X_d_1 = X_d_1
        self.T_f_1 = T_f_1
        self.X_f_2 = X_f_2
        self.X_d_2 = X_d_2
        self.T_f_2 = T_f_2
        self.X_f_3 = X_f_3
        self.X_d_3 = X_d_3
        self.T_f_3 = T_f_3
        self.B_psi = B_psi
        # self.H1 = H1
        # self.H2 = H2
        # self.H3 = H3
        self.H = H
        
        self.psi_11_burn = None
        self.psi_12_burn = None
        self.psi_13_burn = None
        self.psi_21_burn = None
        self.psi_22_burn = None
        self.psi_23_burn = None
        self.psi_31_burn = None
        self.psi_32_burn = None
        self.psi_33_burn = None   
        
        self.psi_11_normal = None
        self.psi_12_normal = None
        self.psi_13_normal = None
        self.psi_21_normal = None
        self.psi_22_normal = None
        self.psi_23_normal = None
        self.psi_31_normal = None
        self.psi_32_normal = None
        self.psi_33_normal = None

        self.beta1_burn = None
        self.beta2_burn = None
        self.beta3_burn = None
        self.beta1_normal = None
        self.beta2_normal = None
        self.beta3_normal = None
        
        self.gamma_0_burn = None
        self.gamma_0_normal = None
         
        self.gamma_1_burn = None
        self.gamma_1_normal = None
                
        self.sigma_e_2_burn = None        
        self.sigma_e_2_normal = None
        self.nt = None
        self.num_func = 3 #PDE
        
        self.psi11_set = []
        self.psi12_set = []
        self.psi13_set = []
        self.psi21_set = []
        self.psi22_set = []
        self.psi23_set = []
        self.psi31_set = []
        self.psi32_set = []
        self.psi33_set = []
        
        self.sigma_set = []
        self.gamma0_set = []
        self.gamma1_set = []
        
        self.beta1_set = []
        self.beta2_set = []
        self.beta3_set = []

    @staticmethod
    def SSE_condition(Y_curr, B_curr, beta_curr):
        return (Y_curr - B_curr@beta_curr).T@(Y_curr - B_curr@beta_curr)

    def psi_mat(self, t_cur):
        temp = self.B_psi@t_cur
        con = np.array([temp for i in range(self.nt)])
        return con.reshape(self.nt*(temp.shape[0]))

    @staticmethod
    def gamma0_post(a_0, b_0, K, zeta_curr):#gamma postertior
        return gamma(a = a_0 + K/2, scale = 1/(b_0 + zeta_curr/(2*3)))
    
    def gamma1_post(self, a_1, b_1, K, beta1, beta2, beta3):
        return gamma(a = a_1 + K/2, scale = 1/(b_1 + (beta1.T@(self.H)@beta1 + beta2.T@(self.H)@beta2 + beta3.T@(self.H)@beta3)/6))

    @staticmethod
    def sigma_post(a_e, b_e, n, SSE_curr):#sigma posterior
        return invgamma(a=a_e + n/2, scale=b_e + SSE_curr/(2*3))
    
    def sigma_2(self, res_sigma, a_e, b_e, n, SSE_new):#sigma_2
        states = []
        sigma_dist = self.sigma_post(a_e, b_e, n, SSE_new)
        cur = res_sigma[-1] # current
        next1 = sigma_dist.rvs() #new
        self.sigma_set.append(next1)
        exp1 = sigma_dist.pdf(next1)
        exp2 = sigma_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]     

    def gamma_0(self, res_gamma0, a_0, b_0, K, zeta_curr):#gamma_0
        states = []
        gamma_dist = self.gamma0_post(a_0, b_0, K, zeta_curr)
        cur =  res_gamma0[-1] # current
        next1 = gamma_dist.rvs() #new
        self.gamma0_set.append(next1)
        exp1 = gamma_dist.pdf(next1)
        exp2 = gamma_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]    
    
    def gamma_1(self, res_gamma1, a_1, b_1, K, beta1, beta2, beta3):#gamma_1
        states = []
        gamma1_dist = self.gamma1_post(a_1, b_1, K, beta1, beta2, beta3)
        cur =  res_gamma1[-1] # current
        next1 = gamma1_dist.rvs() #new
        self.gamma1_set.append(next1)
        exp1 = gamma1_dist.pdf(next1)
        exp2 = gamma1_dist.pdf(cur)
        if exp1 >= exp2:
            states.append(next1)
        else:
            states.append(cur)
        return states[-1]

    def F_cal(self, T_f, X_d, X_f, B, psi_1_curr, psi_2_curr, psi_3_curr):
        return T_f - np.diag(self.psi_mat(psi_1_curr))@ X_d - np.diag(self.psi_mat(psi_2_curr))@X_f - np.diag(self.psi_mat(psi_3_curr))@B

    def D_cal(self, B, sigma_curr, gamma0_curr, gamma1_curr, F_curr):
        return (np.linalg.inv(B.T@B + sigma_curr*(gamma0_curr*F_curr.T@F_curr + gamma1_curr*self.H)))

    @staticmethod
    def beta_post(D, B, Y, sigma):#sigma posterior
        return multivariate_normal(mean=D@B.T@Y, cov=sigma*D)

    def Beta_sample(self, res_beta, D, B, Y, sigma):#beta
        beta_dist = self.beta_post(D, B, Y, sigma)
        cur =  res_beta[-1] # current
        next1 = beta_dist.rvs() #new

        # exp1 = beta_dist.pdf(next1)
        # exp2 = beta_dist.pdf(cur)
        exp1 = -(next1.T@np.linalg.inv(D)@next1 - 2*next1.T@B.T@Y)/(2*sigma)
        exp2 = -(cur.T@np.linalg.inv(D)@cur - 2*cur.T@B.T@Y)/(2*sigma)
        if exp2 < exp1:
            return next1
        else:
            return cur
    
    @staticmethod
    def psi_single(psi_i, sigma_prop):#psi posterior
        return norm(loc=psi_i, scale=sigma_prop)
    
    def zeta_cal(self, tm1, tm2, tm3, T_f, X_d, X_f, beta):
        tm1_mat = self.psi_mat(tm1)
        tm2_mat = self.psi_mat(tm2)
        tm3_mat = self.psi_mat(tm3)
        return T_f@beta - np.multiply(tm1_mat, X_d@beta) - np.multiply(tm2_mat, X_f@beta) - np.multiply(tm3_mat, self.B_mat@beta)
    
    @staticmethod
    def p_cal(a, b):
        pp = 1 - 0.01*(pearsonr(a, b)[0])
        return pp
    
    def psi_1(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_t31, res_t32, res_t33, res_beta1, res_beta2, res_beta3, sigma_prop, gamma_0):#psi_x1
        states1 = []
        states2 = []
        states3 = []

        t11_cur = res_t11 # current
        print('ψ_11:', t11_cur)
        t12_cur = res_t12
        t13_cur = res_t13
        
        t21_cur = res_t21
        print('ψ_21:', t21_cur)
        t22_cur = res_t22
        t23_cur = res_t23
        t31_cur = res_t31
        print('ψ_31:', t31_cur)
        t32_cur = res_t32
        t33_cur = res_t33        
        
        SS = np.array([[1,0.9,0.9], [0.9,1,0.9],[0.9,0.9,1]])#new
        # SS = np.array([[3,0,0], [0,3,0],[0,0,3]])#new
        INV_SIGMA = np.linalg.pinv(SS)#new
        tx2 = np.array([t12_cur, t22_cur, t32_cur])#new
        tx3 = np.array([t13_cur, t23_cur, t33_cur])#new
                
        shape = len(t11_cur)
        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        new_psi1 = []
        new_psi2 = []
        new_psi3 = []
        for i in range(shape):#t11
            new1 = (self.psi_single(t11_cur[i], sigma_prop).rvs()) #new
            next_t11 = copy.deepcopy(t11_cur)
            next_t11[i] = new1
            new_psi1.append(new1)
            
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new1 = self.zeta_cal(next_t11, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            
            p12old = self.p_cal(t11_cur, t21_cur)
            p12new = self.p_cal(next_t11, t21_cur)
            p13old = self.p_cal(t11_cur, t31_cur)
            p13new = self.p_cal(next_t11, t31_cur)
            
            zeta_sumsq_old1.append(zeta_old1.T@zeta_old1 + p12old*zeta_old2.T@zeta_old2 + p13old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new1.append(zeta_new1.T@zeta_new1 + p12new*zeta_old2.T@zeta_old2 + p13new*zeta_old3.T@zeta_old3)
            
            # zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            # zeta_origin_new = zeta_new1.T@zeta_new1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            
            tx1 = np.array([t11_cur, t21_cur, t31_cur])
            tx1_new = np.array([next_t11, t21_cur, t31_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1_new.T@INV_SIGMA@tx1_new) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            # print('sqcur', square_cur)
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old1[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new1[-1]/2)

            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)

            if exp2 > exp1:
                states1.append(next_t11)
                t11_cur = next_t11
            else:
                states1.append(t11_cur)
        
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []
        
        for i in range(shape):#t21
            new2 = (self.psi_single(t21_cur[i], sigma_prop).rvs()) #new
            next_t21 = copy.deepcopy(t21_cur)
            next_t21[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new2 = self.zeta_cal(next_t21, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            
            p21old = self.p_cal(t21_cur, t11_cur)
            p21new = self.p_cal(next_t21, t11_cur)
            p23old = self.p_cal(t21_cur, t31_cur)
            p23new = self.p_cal(next_t21, t31_cur)
            
            zeta_sumsq_old2.append(p21old*zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + p23old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new2.append(p21new*zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + p23new*zeta_old3.T@zeta_old3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + zeta_old3.T@zeta_old3
            
            tx1 = np.array([t11_cur, t21_cur, t31_cur])
            tx1_new = np.array([t11_cur, next_t21, t31_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1_new.T@INV_SIGMA@tx1_new) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old2[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new2[-1]/2)

            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)

            if exp2 > exp1:
                states2.append(next_t21)
                t21_cur = next_t21
            else:
                states2.append(t21_cur)
                
        zeta_sumsq_old3 = []
        zeta_sumsq_new3 = []
        for i in range(shape):#t31
            new3 = (self.psi_single(t31_cur[i], sigma_prop).rvs()) #new
            next_t31 = copy.deepcopy(t31_cur)
            next_t31[i] = new3
            new_psi3.append(new3)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new3 = self.zeta_cal(next_t31, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            
            p31old = self.p_cal(t31_cur, t11_cur)
            p31new = self.p_cal(next_t31, t11_cur)
            p32old = self.p_cal(t31_cur, t21_cur)
            p32new = self.p_cal(next_t31, t21_cur)
            
            zeta_sumsq_old3.append(p31old*zeta_old1.T@zeta_old1 + p32old*zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3)
            zeta_sumsq_new3.append(p31new*zeta_old1.T@zeta_old1 + p32new*zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3
            
            tx1 = np.array([t11_cur, t21_cur, t31_cur])
            tx1_new = np.array([t11_cur, t21_cur, next_t31])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1_new.T@INV_SIGMA@tx1_new) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old3[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new3[-1]/2)
            # 
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)
            
            if exp2 > exp1:
                states3.append(next_t31)
                t31_cur = next_t31
            else:
                states3.append(t31_cur)
        self.psi11_set.append(new_psi1)
        self.psi21_set.append(new_psi2)
        self.psi31_set.append(new_psi3)
        print('p*zeta:',zeta_sumsq_old3[-1], 'p*zeta new:',zeta_sumsq_new3[-1])
        return states1[-1], states2[-1], states3[-1]
    
    def psi_2(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_t31, res_t32, res_t33, res_beta1, res_beta2, res_beta3, sigma_prop, gamma_0):#抽样psi_x2
        states1 = []
        states2 = []
        states3 = []

        t11_cur = res_t11
        t12_cur = res_t12 #current
        t13_cur = res_t13
        print('ψ_12:', t12_cur)
        
        t21_cur = res_t21
        t22_cur = res_t22
        print('ψ_22:', t22_cur)
        t23_cur = res_t23
        t31_cur = res_t31
        t32_cur = res_t32
        print('ψ_32:', t32_cur)
        t33_cur = res_t33        

        SS = np.array([[1,0.9,0.9], [0.9,1,0.9],[0.9,0.9,1]])#new
        INV_SIGMA = np.linalg.pinv(SS)#new
        tx1 = np.array([t11_cur, t21_cur, t31_cur])#new
        tx3 = np.array([t13_cur, t23_cur, t33_cur])#new
        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []
        zeta_sumsq_old3 = []
        zeta_sumsq_new3 = []
        new_psi1 = []
        new_psi2 = []
        new_psi3 = []
        shape = len(t12_cur)
        for i in range(shape):#t12
            new1 = (self.psi_single(t12_cur[i], sigma_prop).rvs()) #new
            next_t12 = copy.deepcopy(t12_cur)
            next_t12[i] = new1
            new_psi1.append(new1)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new1 = self.zeta_cal(t11_cur, next_t12, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            
            p12old = self.p_cal(t12_cur, t22_cur)
            p12new = self.p_cal(next_t12, t22_cur)
            p13old = self.p_cal(t12_cur, t32_cur)
            p13new = self.p_cal(next_t12, t32_cur)

            zeta_sumsq_old1.append(zeta_old1.T@zeta_old1 + p12old*zeta_old2.T@zeta_old2 + p13old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new1.append(zeta_new1.T@zeta_new1 + p12new*zeta_old2.T@zeta_old2 + p13new*zeta_old3.T@zeta_old3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_new1.T@zeta_new1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            
            tx2 = np.array([t12_cur, t22_cur, t32_cur])
            tx2_new = np.array([next_t12, t22_cur, t32_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2_new.T@INV_SIGMA@tx2_new) + np.trace(tx3.T@INV_SIGMA@tx3)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old1[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new1[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)
            
            if exp2 > exp1:
                states1.append(next_t12)
                t12_cur = next_t12
            else:
                states1.append(t12_cur)
        
        for i in range(shape):#t22
            new2 = (self.psi_single(t22_cur[i], sigma_prop).rvs()) #new
            next_t22 = copy.deepcopy(t22_cur)
            next_t22[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new2 = self.zeta_cal(t21_cur, next_t22, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)

            p21old = self.p_cal(t22_cur, t12_cur)
            p21new = self.p_cal(next_t22, t12_cur)
            p23old = self.p_cal(t22_cur, t32_cur)
            p23new = self.p_cal(next_t22, t32_cur)

            zeta_sumsq_old2.append(p21old*zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + p23old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new2.append(p21new*zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + p23new*zeta_old3.T@zeta_old3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + zeta_old3.T@zeta_old3
            
            tx2 = np.array([t12_cur, t22_cur, t32_cur])
            tx2_new = np.array([t12_cur, next_t22, t32_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2_new.T@INV_SIGMA@tx2_new) + np.trace(tx3.T@INV_SIGMA@tx3)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old2[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new2[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)
            if exp2 > exp1:
                states2.append(next_t22)
                t22_cur = next_t22
            else:
                states2.append(t22_cur)
        
        for i in range(shape):#t32
            new3 = (self.psi_single(t32_cur[i], sigma_prop).rvs()) #new
            next_t32 = copy.deepcopy(t32_cur)
            next_t32[i] = new3
            new_psi3.append(new3)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new3 = self.zeta_cal(t31_cur, next_t32, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)

            p31old = self.p_cal(t32_cur, t12_cur)
            p31new = self.p_cal(next_t32, t12_cur)
            p32old = self.p_cal(t32_cur, t22_cur)
            p32new = self.p_cal(next_t32, t22_cur)

            zeta_sumsq_old3.append(p31old*zeta_old1.T@zeta_old1 + p32old*zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3)
            zeta_sumsq_new3.append(p31new*zeta_old1.T@zeta_old1 + p32new*zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3
            
            tx2 = np.array([t12_cur, t22_cur, t32_cur])
            tx2_new = np.array([t12_cur, t22_cur, next_t32])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2_new.T@INV_SIGMA@tx2_new) + np.trace(tx3.T@INV_SIGMA@tx3)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old3[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new3[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)
            if exp2 > exp1:
                states3.append(next_t32)
                t32_cur = next_t32
            else:
                states3.append(t32_cur)
        
        self.psi12_set.append(new_psi1)
        self.psi22_set.append(new_psi2)
        self.psi32_set.append(new_psi3)
        print('p*zeta:',zeta_sumsq_old3[-1], 'p*zeta new:',zeta_sumsq_new3[-1])
        return states1[-1], states2[-1], states3[-1]
    
    def psi_3(self, res_t11, res_t12, res_t13, res_t21, res_t22, res_t23, res_t31, res_t32, res_t33, res_beta1, res_beta2, res_beta3, sigma_prop, gamma_0):#抽样psi_x3
        states1 = []
        states2 = []
        states3 = []

        t11_cur = res_t11
        t12_cur = res_t12 
        t13_cur = res_t13# current
        print('ψ_13:', t13_cur)
        
        t21_cur = res_t21
        t22_cur = res_t22
        t23_cur = res_t23
        print('ψ_23:', t23_cur)
        t31_cur = res_t31
        t32_cur = res_t32
        t33_cur = res_t33
        print('ψ_33:', t33_cur)
        
        SS = np.array([[1,0.9,0.9], [0.9,1,0.9],[0.9,0.9,1]])#new
        INV_SIGMA = np.linalg.pinv(SS)#new
        
        tx1 = np.array([t11_cur, t21_cur, t31_cur])#new
        tx2 = np.array([t12_cur, t22_cur, t32_cur])#new
        
        zeta_sumsq_old1 = []
        zeta_sumsq_new1 = []
        zeta_sumsq_old2 = []
        zeta_sumsq_new2 = []
        zeta_sumsq_old3 = []
        zeta_sumsq_new3 = []
        
        new_psi1 = []
        new_psi2 = []
        new_psi3 = []
        shape = len(t13_cur)
        
        for i in range(shape):#t13
            new1 = (self.psi_single(t13_cur[i], sigma_prop).rvs()) #new
            next_t13 = copy.deepcopy(t13_cur)
            next_t13[i] = new1
            new_psi1.append(new1)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new1 = self.zeta_cal(t11_cur, t12_cur, next_t13, T_f_1, X_d_1, X_f_1, res_beta1)

            p12old = self.p_cal(t13_cur, t23_cur)
            p12new = self.p_cal(next_t13, t23_cur)
            p13old = self.p_cal(t13_cur, t33_cur)
            p13new = self.p_cal(next_t13, t33_cur)

            zeta_sumsq_old1.append(zeta_old1.T@zeta_old1 + p12old*zeta_old2.T@zeta_old2 + p13old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new1.append(zeta_new1.T@zeta_new1 + p12new*zeta_old2.T@zeta_old2 + p13new*zeta_old3.T@zeta_old3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_new1.T@zeta_new1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            
            tx3 = np.array([t13_cur, t23_cur, t33_cur])
            tx3_new = np.array([next_t13, t23_cur, t33_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3_new.T@INV_SIGMA@tx3_new)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old1[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new1[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)
            
            if exp2 > exp1:
                states1.append(next_t13)
                t13_cur = next_t13
            else:
                states1.append(t13_cur)
        
        for i in range(shape):#t23
            new2 = (self.psi_single(t23_cur[i], sigma_prop).rvs()) #new
            next_t23 = copy.deepcopy(t23_cur)
            next_t23[i] = new2
            new_psi2.append(new2)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new2 = self.zeta_cal(t21_cur, t22_cur, next_t23, T_f_2, X_d_2, X_f_2, res_beta2)

            p21old = self.p_cal(t23_cur, t13_cur)
            p21new = self.p_cal(next_t23, t13_cur)
            p23old = self.p_cal(t23_cur, t33_cur)
            p23new = self.p_cal(next_t23, t33_cur)
            
            zeta_sumsq_old2.append(p21old*zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + p23old*zeta_old3.T@zeta_old3)
            zeta_sumsq_new2.append(p21new*zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + p23new*zeta_old3.T@zeta_old3)
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_new2.T@zeta_new2 + zeta_old3.T@zeta_old3
            
            tx3 = np.array([t13_cur, t23_cur, t33_cur])
            tx3_new = np.array([t13_cur, next_t23, t33_cur])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3_new.T@INV_SIGMA@tx3_new)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old2[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new2[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)

            if exp2 > exp1:
                states2.append(next_t23)
                t23_cur = next_t23
            else:
                states2.append(t23_cur)
                
        for i in range(shape):#t33
            new3 = (self.psi_single(t33_cur[i], sigma_prop).rvs()) #new
            next_t33 = copy.deepcopy(t33_cur)
            next_t33[i] = new3
            new_psi3.append(new3)
            zeta_old1 = self.zeta_cal(t11_cur, t12_cur, t13_cur, T_f_1, X_d_1, X_f_1, res_beta1)
            zeta_old2 = self.zeta_cal(t21_cur, t22_cur, t23_cur, T_f_2, X_d_2, X_f_2, res_beta2)
            zeta_old3 = self.zeta_cal(t31_cur, t32_cur, t33_cur, T_f_3, X_d_3, X_f_3, res_beta3)
            zeta_new3 = self.zeta_cal(t31_cur, t32_cur, next_t33, T_f_3, X_d_3, X_f_3, res_beta3)

            p31old = self.p_cal(t33_cur, t13_cur)
            p31new = self.p_cal(next_t33, t13_cur)
            p32old = self.p_cal(t33_cur, t23_cur)
            p32new = self.p_cal(next_t33, t23_cur)

            zeta_sumsq_old3.append(p31old*zeta_old1.T@zeta_old1 + p32old*zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3)
            zeta_sumsq_new3.append(p31new*zeta_old1.T@zeta_old1 + p32new*zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3)
            
            zeta_origin_old = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_old3.T@zeta_old3
            zeta_origin_new = zeta_old1.T@zeta_old1 + zeta_old2.T@zeta_old2 + zeta_new3.T@zeta_new3
            
            tx3 = np.array([t13_cur, t23_cur, t33_cur])
            tx3_new = np.array([t13_cur, t23_cur, next_t33])
            
            square_cur = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3.T@INV_SIGMA@tx3)
            square_new = np.trace(tx1.T@INV_SIGMA@tx1) + np.trace(tx2.T@INV_SIGMA@tx2) + np.trace(tx3_new.T@INV_SIGMA@tx3_new)
            
            exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_old3[-1]/2)
            exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_sumsq_new3[-1]/2)
            # exp1 = (-square_cur/(2*sigma_psi*shape) - gamma_0*zeta_origin_old/2)
            # exp2 = (-square_new/(2*sigma_psi*shape) - gamma_0*zeta_origin_new/2)

            if exp2 > exp1:
                states3.append(next_t33)
                t33_cur = next_t33
            else:
                states3.append(t33_cur)
        print('p*zeta:',zeta_sumsq_old3[-1], 'p*zeta new:',zeta_sumsq_new3[-1])
        self.psi13_set.append(new_psi1)
        self.psi23_set.append(new_psi2)
        self.psi33_set.append(new_psi3)
        return states1[-1], states2[-1], states3[-1]   
    
    def gibbs_burn(self, N_burn, n, K, a_e, b_e, a_0, b_0, a_1, b_1, sigma_prop, nt, shape):
        res_sigma = []
        res_gamma0 = []
        res_gamma1 = []
        
        res_psi11 = []#PDE1 2nd derivative
        res_psi12 = []#PDE1 1st derivative
        res_psi13 = []#PDE1 constant
        res_psi21 = []#PDE2 2nd derivative
        res_psi22 = []#PDE2 1st derivative
        res_psi23 = []#PDE2 constant
        res_psi31 = []#PDE3 2nd derivative
        res_psi32 = []#PDE3 1st derivative
        res_psi33 = []#PDE3 constant
        
        res_beta1 = []
        res_beta2 = []
        res_beta3 = []
        
        self.nt = nt
        zeta_burn = []
        
        # 1.Bspline error
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.Beta1)
        SSE_new2 = self.SSE_condition(self.Y2_obs, self.B_mat, self.Beta2)
        SSE_new3 = self.SSE_condition(self.Y3_obs, self.B_mat, self.Beta3)
        SSE_new = SSE_new1 + SSE_new2 + SSE_new3
        print('SSE origin', SSE_new)
        
        # 2.时time-varying parameter initilization
        psi_11_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_12_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_13_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))

        psi_21_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_22_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_23_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))

        psi_31_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_32_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        psi_33_curr = (np.random.multivariate_normal(np.zeros([shape]), 1*np.eye(shape)))
        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        res_psi21.append(psi_21_curr)
        res_psi22.append(psi_22_curr)
        res_psi23.append(psi_23_curr)
        res_psi31.append(psi_31_curr)
        res_psi32.append(psi_32_curr)
        res_psi33.append(psi_33_curr)
        
        # 3.sigma initialization
        res_sigma.append((self.sigma_post(a_e, b_e, n, SSE_new)).rvs())
        
        # 4.beta initializaiton
        res_beta1.append(self.Beta1)
        res_beta2.append(self.Beta2)
        res_beta3.append(self.Beta3)
        
        # 5.PDE error initializaiton
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, T_f_1, X_d_1, X_f_1, self.Beta1) #zeta initializaiton
        zeta_new2 = self.zeta_cal(psi_21_curr, psi_22_curr, psi_23_curr, T_f_2, X_d_2, X_f_2, self.Beta2) #zeta initializaiton
        zeta_new3 = self.zeta_cal(psi_31_curr, psi_32_curr, psi_33_curr, T_f_3, X_d_3, X_f_3, self.Beta3) #zeta initializaiton
        zeta_new = zeta_new1.T@zeta_new1 + zeta_new2.T@zeta_new2 + zeta_new3.T@zeta_new3
        print('zeta origin', zeta_new)
        zeta_burn.append(zeta_new)
        
        # 6.gamma initializaiton
        res_gamma0.append(self.gamma0_post(a_0, b_0, K, zeta_new).rvs())#gamma0 initializaiton
        res_gamma1.append(self.gamma1_post(a_1, b_1, K, res_beta1[-1],  res_beta2[-1],  res_beta3[-1]).rvs())#gamma1 initializaiton
        
        # iteration
        for i in range(N_burn):
            # 7.update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))
            
            # 8.update amma
            res_gamma0.append(self.gamma_0(res_gamma0, a_0, b_0, K, zeta_burn[-1]))
            res_gamma1.append(self.gamma_1(res_gamma1, a_1, b_1, K, res_beta1[-1],  res_beta2[-1],  res_beta3[-1]))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma0[-1])
            print('Current gamma1:', res_gamma1[-1])
            
            # 9. update beta
            F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.X_f_1, self.B_mat, res_psi11[-1], res_psi12[-1], res_psi13[-1])
            F_new2 = self.F_cal(self.T_f_2, self.X_d_2, self.X_f_2, self.B_mat, res_psi21[-1], res_psi22[-1], res_psi23[-1])
            F_new3 = self.F_cal(self.T_f_3, self.X_d_3, self.X_f_3, self.B_mat, res_psi31[-1], res_psi32[-1], res_psi33[-1])
            D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new1)
            D_new2 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new2)
            D_new3 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new3)
            res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            res_beta2.append(self.Beta_sample(res_beta2, D_new2, self.B_mat, self.Y2_obs, res_sigma[-1]))
            res_beta3.append(self.Beta_sample(res_beta3, D_new3, self.B_mat, self.Y3_obs, res_sigma[-1]))
            
            # 10.update time-varying parameters ψ
            sample11, sample21, sample31 = self.psi_1(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi11.append(sample11)
            res_psi21.append(sample21)
            res_psi31.append(sample31)            
            sample12, sample22, sample32 = self.psi_2(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi12.append(sample12)
            res_psi22.append(sample22)
            res_psi32.append(sample32)            
            sample13, sample23, sample33 = self.psi_3(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi13.append(sample13)
            res_psi23.append(sample23)
            res_psi33.append(sample33)

            # 11.update Bspline errors
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_2 = self.SSE_condition(self.Y2_obs, self.B_mat, res_beta2[-1])
            SSE_3 = self.SSE_condition(self.Y3_obs, self.B_mat, res_beta3[-1])
            SSE_new = SSE_1 + SSE_2 + SSE_3
            
            # 12.update PDE误差
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], T_f_1, X_d_1, X_f_1, res_beta1[-1])
            zeta_2 = self.zeta_cal(res_psi21[-1], res_psi22[-1], res_psi23[-1], T_f_2, X_d_2, X_f_2, res_beta2[-1])
            zeta_3 = self.zeta_cal(res_psi31[-1], res_psi32[-1], res_psi33[-1], T_f_3, X_d_3, X_f_3, res_beta3[-1])
            zeta_new = zeta_1.T@zeta_1 + zeta_2.T@zeta_2 + zeta_3.T@zeta_3   
            zeta_burn.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new, 'current_BSpline_SST:', SSE_new)
            print('------------------')
        
        self.sigma_e_2_burn = res_sigma[-1]
        self.gamma_0_burn = res_gamma0[-1]
        self.gamma_1_burn = res_gamma1[-1]
        self.beta1_burn = res_beta1[-1]
        self.beta2_burn = res_beta2[-1]
        self.beta3_burn = res_beta3[-1]
        self.psi_11_burn = res_psi11[-1]
        self.psi_12_burn = res_psi12[-1]
        self.psi_13_burn = res_psi13[-1]
        self.psi_21_burn = res_psi21[-1]
        self.psi_22_burn = res_psi22[-1]
        self.psi_23_burn = res_psi23[-1]
        self.psi_31_burn = res_psi31[-1]
        self.psi_32_burn = res_psi32[-1]
        self.psi_33_burn = res_psi33[-1]

        print("sigma_e_2 estimation:",res_sigma[-1])
        print("gamma_0 estimation:",res_gamma0[-1])
        print("gamma_1 estimation:",res_gamma1[-1])
        print("beta1 estimation:",res_beta1[-1])
        print("beta2 estimation:",res_beta2[-1])
        print("beta3 estimation:",res_beta3[-1])
        
        print("psi_11 estimation:",res_psi11[-1])
        print("psi_12 estimation:",res_psi12[-1])
        print("psi_13 estimation:",res_psi13[-1])
        
        print("psi_21 estimation:",res_psi21[-1])
        print("psi_22 estimation:",res_psi22[-1])
        print("psi_23 estimation:",res_psi23[-1])
        
        print("psi_31 estimation:",res_psi31[-1])
        print("psi_32 estimation:",res_psi32[-1])
        print("psi_33 estimation:",res_psi33[-1])
        return res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_psi31, res_psi32, res_psi33, res_beta1, res_beta2, res_beta3, res_sigma, res_gamma0, res_gamma1, zeta_burn
    
    def gibbs_normal(self, N_normal, n, K, a_e, b_e, a_0, b_0, a_1, b_1, sigma_prop, nt_0):
        res_sigma = []
        res_gamma0 = []
        res_gamma1 = []
        
        res_psi11 = []#PDE1 2nd derivative
        res_psi12 = []#PDE1 1st derivative
        res_psi13 = []#PDE1 constant
        res_psi21 = []#PDE2 2nd derivative
        res_psi22 = []#PDE2 1st derivative
        res_psi23 = []#PDE2 constant
        res_psi31 = []#PDE3 2nd derivative
        res_psi32 = []#PDE3 1st derivative
        res_psi33 = []#PDE3 constant
        
        res_beta1 = []
        res_beta2 = []
        res_beta3 = []

        self.nt = nt_0
        zeta_normal = [] # zeta²
        
        # 1.Bspline error
        SSE_new1 = self.SSE_condition(self.Y1_obs, self.B_mat, self.beta1_burn)
        SSE_new2 = self.SSE_condition(self.Y2_obs, self.B_mat, self.beta2_burn)
        SSE_new3 = self.SSE_condition(self.Y3_obs, self.B_mat, self.beta3_burn)
        SSE_new = SSE_new1 + SSE_new2 + SSE_new3
        
        # 2.time-varying parameters
        psi_11_curr = self.psi_11_burn
        psi_12_curr = self.psi_12_burn
        psi_13_curr = self.psi_13_burn

        psi_21_curr = self.psi_21_burn
        psi_22_curr = self.psi_22_burn
        psi_23_curr = self.psi_23_burn

        psi_31_curr = self.psi_31_burn
        psi_32_curr = self.psi_32_burn
        psi_33_curr = self.psi_33_burn
        
        res_psi11.append(psi_11_curr)
        res_psi12.append(psi_12_curr)
        res_psi13.append(psi_13_curr)
        res_psi21.append(psi_21_curr)
        res_psi22.append(psi_22_curr)
        res_psi23.append(psi_23_curr)
        res_psi31.append(psi_31_curr)
        res_psi32.append(psi_32_curr)
        res_psi33.append(psi_33_curr)
        
        # 3.sigma
        res_sigma.append(self.sigma_e_2_burn)
        
         # 4.beta
        res_beta1.append(self.beta1_burn)
        res_beta2.append(self.beta2_burn)
        res_beta3.append(self.beta3_burn)
        
        # 5.PDE errors
        zeta_new1 = self.zeta_cal(psi_11_curr, psi_12_curr, psi_13_curr, T_f_1, X_d_1, X_f_1, res_beta1[-1])
        zeta_new2 = self.zeta_cal(psi_21_curr, psi_22_curr, psi_23_curr, T_f_2, X_d_2, X_f_2, res_beta2[-1])
        zeta_new3 = self.zeta_cal(psi_31_curr, psi_32_curr, psi_33_curr, T_f_3, X_d_3, X_f_3, res_beta3[-1])
        zeta_new = zeta_new1.T@zeta_new1 + zeta_new2.T@zeta_new2 + zeta_new3.T@zeta_new3 
        zeta_normal.append(zeta_new)
        
        # 6.gamm
        res_gamma0.append(self.gamma_0_burn)#gamma0
        res_gamma1.append(self.gamma_1_burn)#gamma1               

        for i in range(N_normal):
            # 7.update sigma
            res_sigma.append(self.sigma_2(res_sigma, a_e, b_e, n, SSE_new))
            
            # 8.update gamma
            res_gamma0.append(self.gamma_0(res_gamma0, a_0, b_0, K, zeta_new))
            res_gamma1.append(self.gamma_1(res_gamma1, a_1, b_1, K, res_beta1[-1], res_beta2[-1], res_beta3[-1]))
            print('Current sigma2:', res_sigma[-1], 'Current gamma:', res_gamma0[-1])
            print('Current gamma1:', res_gamma1[-1])
            
            F_new1 = self.F_cal(self.T_f_1, self.X_d_1, self.X_f_1, self.B_mat, res_psi11[-1], res_psi12[-1], res_psi13[-1])
            F_new2 = self.F_cal(self.T_f_2, self.X_d_2, self.X_f_2, self.B_mat, res_psi21[-1], res_psi22[-1], res_psi23[-1])
            F_new3 = self.F_cal(self.T_f_3, self.X_d_3, self.X_f_3, self.B_mat, res_psi31[-1], res_psi32[-1], res_psi33[-1])
            D_new1 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new1)
            D_new2 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new2)
            D_new3 = self.D_cal(self.B_mat, res_sigma[-1], res_gamma0[-1], res_gamma1[-1], F_new3)
            # 9. update beta
            res_beta1.append(self.Beta_sample(res_beta1, D_new1, self.B_mat, self.Y1_obs, res_sigma[-1]))
            res_beta2.append(self.Beta_sample(res_beta2, D_new2, self.B_mat, self.Y2_obs, res_sigma[-1]))
            res_beta3.append(self.Beta_sample(res_beta3, D_new3, self.B_mat, self.Y3_obs, res_sigma[-1]))
            # res_beta1.append(self.beta1_burn)
            # res_beta2.append(self.beta2_burn)
            # res_beta3.append(self.beta3_burn)
            
            # 10.update time-varying ψ
            sample11, sample21, sample31 = self.psi_1(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi11.append(sample11)
            res_psi21.append(sample21)
            res_psi31.append(sample31)
            
            sample12, sample22, sample32 = self.psi_2(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi12.append(sample12)
            res_psi22.append(sample22)
            res_psi32.append(sample32)
            
            sample13, sample23, sample33 = self.psi_3(res_psi11[-1], res_psi12[-1], res_psi13[-1], res_psi21[-1], res_psi22[-1], res_psi23[-1], res_psi31[-1], res_psi32[-1], res_psi33[-1], self.Beta1, self.Beta2, self.Beta3, sigma_prop, res_gamma0[-1])
            res_psi13.append(sample13)
            res_psi23.append(sample23)
            res_psi33.append(sample33)
            
            # 11.update Bspline errors
            SSE_1 = self.SSE_condition(self.Y1_obs, self.B_mat, res_beta1[-1])
            SSE_2 = self.SSE_condition(self.Y2_obs, self.B_mat, res_beta2[-1])
            SSE_3 = self.SSE_condition(self.Y3_obs, self.B_mat, res_beta3[-1])
            SSE_new = SSE_1 + SSE_2 + SSE_3
            
            # 12.update PDE errors
            zeta_1 = self.zeta_cal(res_psi11[-1], res_psi12[-1], res_psi13[-1], T_f_1, X_d_1, X_f_1, res_beta1[-1])
            zeta_2 = self.zeta_cal(res_psi21[-1], res_psi22[-1], res_psi23[-1], T_f_2, X_d_2, X_f_2, res_beta2[-1])
            zeta_3 = self.zeta_cal(res_psi31[-1], res_psi32[-1], res_psi33[-1], T_f_3, X_d_3, X_f_3, res_beta3[-1])
            zeta_new = zeta_1.T@zeta_1 + zeta_2.T@zeta_2 + zeta_3.T@zeta_3    
            zeta_normal.append(zeta_new)
            print('Iteration', i, 'current_zeta_ssq:', zeta_new, 'current_BSpline_SST:', SSE_new)
            print('------------------')
        
        self.sigma_e_2_normal = res_sigma[-1]
        self.gamma_0_normal = res_gamma0[-1]
        self.gamma_1_normal = res_gamma1[-1]
        self.beta1_normal = res_beta1[-1]
        self.beta2_normal = res_beta2[-1]
        self.beta3_normal = res_beta3[-1]
        self.psi_11_normal = res_psi11[-1]
        self.psi_12_normal = res_psi12[-1]
        self.psi_13_normal = res_psi13[-1]
        self.psi_21_normal = res_psi21[-1]
        self.psi_22_normal = res_psi22[-1]
        self.psi_23_normal = res_psi23[-1]
        self.psi_31_normal = res_psi31[-1]
        self.psi_32_normal = res_psi32[-1]
        self.psi_33_normal = res_psi33[-1]

        print("sigma_e_2 estimation:",res_sigma[-1])
        print("gamma_0 estimation:",res_gamma0[-1])
        print("gamma_1 estimation:",res_gamma1[-1])
        print("beta1 estimation:",res_beta1[-1])
        print("beta2 estimation:",res_beta2[-1])
        print("beta3 estimation:",res_beta3[-1])
        
        print("psi_11 estimation:",res_psi11[-1])
        print("psi_12 estimation:",res_psi12[-1])
        print("psi_13 estimation:",res_psi13[-1])
        
        print("psi_21 estimation:",res_psi21[-1])
        print("psi_22 estimation:",res_psi22[-1])
        print("psi_23 estimation:",res_psi23[-1])
        
        print("psi_31 estimation:",res_psi31[-1])
        print("psi_32 estimation:",res_psi32[-1])
        print("psi_33 estimation:",res_psi33[-1])
        return res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_psi31, res_psi32, res_psi33, res_beta1, res_beta2, res_beta3, res_sigma, res_gamma0, res_gamma1, zeta_normal

#true value
Y1 = np.load('input_simulation/Y_TV_1.npy')
B1 = np.load('input_simulation/B_mat_TV_1.npy')

beta1_ini = np.load('input_simulation/beta_TV_1.npy')
Y1_hat = B1@beta1_ini
X_f_1 = np.load('input_simulation/X_f_TV_1.npy')
X_d_1 = np.load('input_simulation/X_d_TV_1.npy')
T_f_1 = np.load('input_simulation/T_f_TV_1.npy')

Y2 = np.load('input_simulation/Y_TV_2.npy')
B2 = np.load('input_simulation/B_mat_TV_2.npy')

beta2_ini = np.load('input_simulation/beta_TV_2.npy')
Y2_hat = B2@beta2_ini
X_f_2 = np.load('input_simulation/X_f_TV_2.npy')
X_d_2 = np.load('input_simulation/X_d_TV_2.npy')
T_f_2 = np.load('input_simulation/T_f_TV_2.npy')

Y3 = np.load('input_simulation/Y_TV_3.npy')
B3 = np.load('input_simulation/B_mat_TV_3.npy')

beta3_ini = np.load('input_simulation/beta_TV_3.npy')
Y3_hat = B3@beta3_ini
X_f_3 = np.load('input_simulation/X_f_TV_3.npy')
X_d_3 = np.load('input_simulation/X_d_TV_3.npy')
T_f_3 = np.load('input_simulation/T_f_TV_3.npy')

B_psi = np.load('input_simulation/B_psi.npy')
psi_node = np.load('input_simulation/psi_node.npy')

H11 = np.load('input_simulation/H11.npy')
H12 = np.load('input_simulation/H12.npy')
H13 = np.load('input_simulation/H13.npy')
H21 = np.load('input_simulation/H21.npy')
H22 = np.load('input_simulation/H22.npy')
H23 = np.load('input_simulation/H23.npy')
H31 = np.load('input_simulation/H31.npy')
H32 = np.load('input_simulation/H32.npy')
H33 = np.load('input_simulation/H33.npy')

H1 = 1e-3*H11 + 1e-6*H12
H2 = 1e-3*H21 + 1e-6*H22
H3 = 1e-3*H31 + 1e-6*H32
n = B1.shape[0] # num of observation
K = beta1_ini.shape[0] #num of basis functions

order = 4
order_penalty = 3
lambda_penalty = 1e-3
sigma_noise = 0

a_e = 1.01
b_e = 1e-6

a_0 = 1e-7
b_0 = 1e-7

a_1 = 1e-7
b_1 = 1e-7

# a_2 = 1e-5
# b_2 = 1e-5
sigma_psi = 3#1

sigma_prop1 = 0.01
sigma_prop2 = 0.1

nt = 33
shape_psi = B_psi.shape[1] 
N_burn = 10000
N_normal = 5000

# Sampler Definition
GibbsObj_1 = GibbsSampler(Y1, Y2, Y3, B1, beta1_ini, beta2_ini, beta3_ini, X_f_1, X_d_1, T_f_1, X_f_2, X_d_2, T_f_2, X_f_3, X_d_3, T_f_3, B_psi, H1)

# Sampling burn-in stage
start_time1 = perf_counter()
res_psi11, res_psi12, res_psi13, res_psi21, res_psi22, res_psi23, res_psi31, res_psi32, res_psi33, res_beta1, res_beta2, res_beta3, res_sigma, res_gamma0, res_gamma1, zeta_burn = GibbsObj_1.gibbs_burn(N_burn, n, K, a_e, b_e, a_0, b_0, a_1, b_1, sigma_prop1, nt, shape_psi)
end_time1 = perf_counter()
print("Time:", end_time1 - start_time1, "秒")

# result 
np.save('output_3in1_simulation/psi_11_burn.npy',res_psi11)
np.save('output_3in1_simulation/psi_12_burn.npy',res_psi12)
np.save('output_3in1_simulation/psi_13_burn.npy',res_psi13)
np.save('output_3in1_simulation/beta1_burn.npy',res_beta1)

np.save('output_3in1_simulation/psi_21_burn.npy',res_psi21)
np.save('output_3in1_simulation/psi_22_burn.npy',res_psi22)
np.save('output_3in1_simulation/psi_23_burn.npy',res_psi23)
np.save('output_3in1_simulation/beta2_burn.npy',res_beta2)

np.save('output_3in1_simulation/psi_31_burn.npy',res_psi31)
np.save('output_3in1_simulation/psi_32_burn.npy',res_psi32)
np.save('output_3in1_simulation/psi_33_burn.npy',res_psi33)
np.save('output_3in1_simulation/beta3_burn.npy',res_beta3)

np.save('output_3in1_simulation/sigma_burn.npy',res_sigma)
np.save('output_3in1_simulation/gamma0_burn.npy',res_gamma0)
np.save('output_3in1_simulation/gamma1_burn.npy',res_gamma1)

# np.save('output_3in1_simulation/psi_11_steady.npy',res_psi11_normal)
# np.save('output_3in1_simulation/psi_12_steady.npy',res_psi12_normal)
# np.save('output_3in1_simulation/psi_13_steady.npy',res_psi13_normal)
# np.save('output_3in1_simulation/beta1_steady.npy',res_beta1_normal)
# 
# np.save('output_3in1_simulation/psi_21_steady.npy',res_psi21_normal)
# np.save('output_3in1_simulation/psi_22_steady.npy',res_psi22_normal)
# np.save('output_3in1_simulation/psi_23_steady.npy',res_psi23_normal)
# np.save('output_3in1_simulation/beta2_steady.npy',res_beta2_normal)
# 
# np.save('output_3in1_simulation/psi_31_steady.npy',res_psi31_normal)
# np.save('output_3in1_simulation/psi_32_steady.npy',res_psi32_normal)
# np.save('output_3in1_simulation/psi_33_steady.npy',res_psi33_normal)
# np.save('output_3in1_simulation/beta3_steady.npy',res_beta3_normal)
# 
# np.save('output_3in1_simulation/sigma_steady.npy',res_sigma_normal)
# np.save('output_3in1_simulation/gamma0_steady.npy',res_gamma0_normal)
# np.save('output_3in1_simulation/gamma1_steady.npy',res_gamma1_normal)

np.save('output_3in1_simulation/zeta_burn.npy',zeta_burn)
# np.save('output_3in1_simulation/zeta_steady.npy',zeta_normal)

# result
np.save('output_3in1_simulation/sigma_sample.npy',GibbsObj_1.sigma_set)
np.save('output_3in1_simulation/gamma0_sample.npy',GibbsObj_1.gamma0_set)
np.save('output_3in1_simulation/gamma1_sample.npy',GibbsObj_1.gamma1_set)
np.save('output_3in1_simulation/psi11_sample.npy',GibbsObj_1.psi11_set)
np.save('output_3in1_simulation/psi12_sample.npy',GibbsObj_1.psi12_set)
np.save('output_3in1_simulation/psi13_sample.npy',GibbsObj_1.psi13_set)

np.save('output_3in1_simulation/psi21_sample.npy',GibbsObj_1.psi21_set)
np.save('output_3in1_simulation/psi22_sample.npy',GibbsObj_1.psi22_set)
np.save('output_3in1_simulation/psi23_sample.npy',GibbsObj_1.psi23_set)

np.save('output_3in1_simulation/psi31_sample.npy',GibbsObj_1.psi31_set)
np.save('output_3in1_simulation/psi32_sample.npy',GibbsObj_1.psi32_set)
np.save('output_3in1_simulation/psi33_sample.npy',GibbsObj_1.psi33_set)

# estimation
psi_11_hat = res_psi11[-1]
psi_12_hat = res_psi12[-1]
psi_13_hat = res_psi13[-1]
psi_21_hat = res_psi21[-1]
psi_22_hat = res_psi22[-1]
psi_23_hat = res_psi23[-1]
psi_31_hat = res_psi31[-1]
psi_32_hat = res_psi32[-1]
psi_33_hat = res_psi33[-1]
beta1_hat = res_beta1[-1]
beta2_hat = res_beta2[-1]
beta3_hat = res_beta3[-1]
sigma_hat = res_sigma[-1]
gamma0_hat = res_gamma0[-1]
gamma1_hat = res_gamma1[-1]

#PDE error 3in1
zeta_ob1 = GibbsObj_1.zeta_cal(psi_11_hat, psi_12_hat, psi_13_hat, T_f_1, X_d_1, X_f_1, beta1_hat)
zeta_ob2 = GibbsObj_1.zeta_cal(psi_21_hat, psi_22_hat, psi_23_hat, T_f_2, X_d_2, X_f_2, beta2_hat)
zeta_ob3 = GibbsObj_1.zeta_cal(psi_31_hat, psi_32_hat, psi_33_hat, T_f_3, X_d_3, X_f_3, beta3_hat)
zeta_ob_ss = zeta_ob1.T@zeta_ob1 + zeta_ob2.T@zeta_ob2 + zeta_ob3.T@zeta_ob3

print('PDE errors:')
print('mean_zeta', 1/3*(np.mean(zeta_ob1) + np.mean(zeta_ob2) + np.mean(zeta_ob3)))
print('mean_abs_zeta', 1/3*(np.mean(np.abs(zeta_ob1)) + np.mean(np.abs(zeta_ob2)) + np.mean(np.abs(zeta_ob3))))
print('max_zeta', np.max([np.max(zeta_ob1), np.max(zeta_ob2), np.max(zeta_ob3)]))
print('min_zeta', np.min([np.min(zeta_ob1), np.min(zeta_ob2), np.min(zeta_ob3)]))
print('rmse_zeta', np.sqrt(zeta_ob_ss/zeta_ob1.shape[0]/3))
print('SST', zeta_ob_ss)

#Bspline error 3in1
error_bs1 = Y1-B1@beta1_hat
error_bs2 = Y2-B2@beta2_hat
error_bs3 = Y3-B3@beta3_hat
error_ss = error_bs1.T@error_bs1 + error_bs2.T@error_bs2 + error_bs3.T@error_bs3

print('B_Spline Errors:')
print('mean_BSpline', 1/3*(np.mean(error_bs1) + np.mean(error_bs2) + np.mean(error_bs3)))
print('mean_abs_BSpline', 1/3*(np.mean(np.abs(error_bs1)) + np.mean(np.abs(error_bs2)) + np.mean(np.abs(error_bs3))))
print('max_BSpline', np.max([np.max(error_bs1), np.max(error_bs2), np.max(error_bs3)]))
print('min_BSpline', np.min([np.min(error_bs1), np.min(error_bs2), np.min(error_bs3)]))
print('rmse_BSpline', np.sqrt(error_ss/error_bs1.shape[0]/3))
print('SST_BSpline', error_ss)

# setting 1 Trace plot
num1 = 1.01
num2 = 0.99
num3 = 2
num4 = 0

N_1 = 5000
# psi 1 legend
psi_11_lengend = [r'$\psi_{11_%d}$'%(i+1) for i in range(shape_psi)]
psi_12_lengend = [r'$\psi_{12_%d}$'%(i+1) for i in range(shape_psi)]
psi_13_lengend = [r'$\psi_{13_%d}$'%(i+1) for i in range(shape_psi)]
# psi 2
psi_21_lengend = [r'$\psi_{21_%d}$'%(i+1) for i in range(shape_psi)]
psi_22_lengend = [r'$\psi_{22_%d}$'%(i+1) for i in range(shape_psi)]
psi_23_lengend = [r'$\psi_{23_%d}$'%(i+1) for i in range(shape_psi)]
# psi 3
psi_31_lengend = [r'$\psi_{31_%d}$'%(i+1) for i in range(shape_psi)]
psi_32_lengend = [r'$\psi_{32_%d}$'%(i+1) for i in range(shape_psi)]
psi_33_lengend = [r'$\psi_{33_%d}$'%(i+1) for i in range(shape_psi)]

# for Excel output
psi11_node = psi_node[0][0]
psi12_node = psi_node[0][1]
psi13_node = psi_node[0][2]
psi21_node = psi_node[1][0]
psi22_node = psi_node[1][1]
psi23_node = psi_node[1][2]
psi31_node = psi_node[2][0]
psi32_node = psi_node[2][1]
psi33_node = psi_node[2][2]

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment #PatternFill
import datetime
now = datetime.datetime.now()

wb = Workbook()
dest_filename = "%s_Time_varying_3in1_simulation_result.xlsx"%(now.strftime('%Y%m%d-%H%M%S'))

ws = wb.active
ws.title = 'Result'

ws['A1'] = "Date"
ws['B1'] = now.strftime('%Y-%m-%d-%H:%M:%S')#%H:%M:%S
ws['C1'] = "Result"
ws.merge_cells(range_string='C1:E1')

ws.append(["psi11", '真值']+[psi11_node[i] for i in range(shape_psi)])
ws.append(["psi11", '采样估计值']+[psi_11_hat[i]for i in range(shape_psi)])
ws.append(["psi11", '误差百分比']+[abs((psi11_node[i]-psi_11_hat[i])/psi11_node[i]) for i in range(shape_psi)])

ws.append(["psi12", '真值']+[psi12_node[i] for i in range(shape_psi)])
ws.append(["psi12", '采样估计值']+[psi_12_hat[i]for i in range(shape_psi)])
ws.append(["psi12", '误差百分比']+[abs((psi12_node[i]-psi_12_hat[i])/psi12_node[i]) for i in range(shape_psi)])

ws.append(["psi13", '真值']+[psi13_node[i] for i in range(shape_psi)])
ws.append(["psi13", '采样估计值']+[psi_13_hat[i]for i in range(shape_psi)])
ws.append(["psi13", '误差百分比']+[abs((psi13_node[i]-psi_13_hat[i])/psi13_node[i]) for i in range(shape_psi)])

ws.append(["psi21", '真值']+[psi21_node[i] for i in range(shape_psi)])
ws.append(["psi21", '采样估计值']+[psi_21_hat[i]for i in range(shape_psi)])
ws.append(["psi21", '误差百分比']+[abs((psi21_node[i]-psi_21_hat[i])/psi21_node[i]) for i in range(shape_psi)])

ws.append(["psi22", '真值']+[psi22_node[i] for i in range(shape_psi)])
ws.append(["psi22", '采样估计值']+[psi_22_hat[i]for i in range(shape_psi)])
ws.append(["psi22", '误差百分比']+[abs((psi22_node[i]-psi_22_hat[i])/psi22_node[i]) for i in range(shape_psi)])

ws.append(["psi23", '真值']+[psi23_node[i] for i in range(shape_psi)])
ws.append(["psi23", '采样估计值']+[psi_23_hat[i]for i in range(shape_psi)])
ws.append(["psi23", '误差百分比']+[abs((psi23_node[i]-psi_23_hat[i])/psi23_node[i]) for i in range(shape_psi)])

ws.append(["psi31", '真值']+[psi11_node[i] for i in range(shape_psi)])
ws.append(["psi31", '采样估计值']+[psi_31_hat[i]for i in range(shape_psi)])
ws.append(["psi31", '误差百分比']+[abs((psi11_node[i]-psi_31_hat[i])/psi11_node[i]) for i in range(shape_psi)])

ws.append(["psi32", '真值']+[psi12_node[i] for i in range(shape_psi)])
ws.append(["psi32", '采样估计值']+[psi_32_hat[i]for i in range(shape_psi)])
ws.append(["psi32", '误差百分比']+[abs((psi12_node[i]-psi_32_hat[i])/psi12_node[i]) for i in range(shape_psi)])

ws.append(["psi33", '真值']+[psi13_node[i] for i in range(shape_psi)])
ws.append(["psi33", '采样估计值']+[psi_33_hat[i]for i in range(shape_psi)])
ws.append(["psi33", '误差百分比']+[abs((psi13_node[i]-psi_33_hat[i])/psi13_node[i]) for i in range(shape_psi)])

ws.append(["sigma_2", '采样估计值', sigma_hat, (b_e+(error_ss)/6)/(a_e+n/2-1)])
ws.append(["gamma0", '采样估计值', gamma0_hat, (a_0+K/2)/(b_0+(zeta_ob_ss)/6)])
ws.append(["gamma1", '采样估计值', gamma1_hat, (a_1+K/2)/(b_1+(beta1_hat.T@(H1)@beta1_hat + beta2_hat.T@(H2)@beta2_hat + beta3_hat.T@(H3)@beta3_hat)/6)])

ws.merge_cells(range_string='A2:A4')
ws.merge_cells(range_string='A5:A7')
ws.merge_cells(range_string='A8:A10')

ws.merge_cells(range_string='A11:A13')
ws.merge_cells(range_string='A14:A16')
ws.merge_cells(range_string='A17:A19')

ws.merge_cells(range_string='A20:A22')
ws.merge_cells(range_string='A23:A25')
ws.merge_cells(range_string='A26:A28')

ws['I1'] = "Model Errors"
ws.merge_cells(range_string='I1:L1')

ws['I2'] = "PDE 1"
ws.merge_cells(range_string='I2:I7')
ws['I14'] = "PDE 2"
ws.merge_cells(range_string='I14:I19')
ws['I26'] = "PDE 3"
ws.merge_cells(range_string='I26:I31')

J = ['mean_zeta', 'mean_abs_zeta','max_zeta','min_zeta','rmse_zeta','SST_zeta', 'mean_BSpline','mean_abs_BSpline','max_BSpline','min_BSpline','rmse_BSpline','SST_BSpline']
K1 = ['均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和','均值', '绝对值均值', '最大值', '最小值', '均方根误差', '总误差平方和']

L1 = [np.mean(zeta_ob1), np.mean(np.abs(zeta_ob1)), np.max(zeta_ob1), np.min(zeta_ob1), np.sqrt(zeta_ob1.T@zeta_ob1/zeta_ob1.shape[0]), zeta_ob1.T@zeta_ob1,np.mean(error_bs1), np.mean(np.abs(error_bs1)), np.max(error_bs1), np.min(error_bs1), np.sqrt(error_bs1.T@error_bs1/error_bs1.shape[0]), error_bs1.T@error_bs1]

L2 = [np.mean(zeta_ob2), np.mean(np.abs(zeta_ob2)), np.max(zeta_ob2), np.min(zeta_ob2), np.sqrt(zeta_ob2.T@zeta_ob2/zeta_ob2.shape[0]), zeta_ob2.T@zeta_ob2,np.mean(error_bs2), np.mean(np.abs(error_bs2)), np.max(error_bs2), np.min(error_bs2), np.sqrt(error_bs2.T@error_bs2/error_bs2.shape[0]), error_bs2.T@error_bs2]

L3 = [np.mean(zeta_ob3), np.mean(np.abs(zeta_ob3)), np.max(zeta_ob3), np.min(zeta_ob3), np.sqrt(zeta_ob3.T@zeta_ob3/zeta_ob3.shape[0]), zeta_ob3.T@zeta_ob3,np.mean(error_bs3), np.mean(np.abs(error_bs3)), np.max(error_bs3), np.min(error_bs3), np.sqrt(error_bs3.T@error_bs3/error_bs3.shape[0]), error_bs3.T@error_bs3]

for i in range(2,14):
    ws['J%d'%i] = J[i-2]
    ws['K%d'%i] = K1[i-2]
    ws['L%d'%i] = L1[i-2]
    
    ws['J%d'%(i+12)] = J[i-2]
    ws['K%d'%(i+12)] = K1[i-2]
    ws['L%d'%(i+12)] = L2[i-2]
    
    ws['J%d'%(i+24)] = J[i-2]
    ws['K%d'%(i+24)] = K1[i-2]
    ws['L%d'%(i+24)] = L3[i-2]
    
ws['I8'] = "统计模型 1"
ws.merge_cells(range_string='I8:I13')
ws['I20'] = "统计模型 2"
ws.merge_cells(range_string='I20:I25')
ws['I32'] = "统计模型 3"
ws.merge_cells(range_string='I32:I37')

ws['K38'] = "PDE_SST"
ws['K39'] = "B_Spline_SST"

ws['L38'] = zeta_ob1.T@zeta_ob1 + zeta_ob2.T@zeta_ob2 + zeta_ob3.T@zeta_ob3
ws['L39'] = error_bs1.T@error_bs1 + error_bs2.T@error_bs2 + error_bs3.T@error_bs3

N = ['#', 'Ver.', 'p_Bspline', 'p_Bpsi', 'n', 'K', 'a_e', 'b_e', 'a_0', 'b_0', 'sigma_psi', 'sigma_prop1', 'sigma_prop2', 'nt', 'N_burn', 'N_normal', 'penalty_order', 'λ']
O = ['Value', 'Time_varying 3in1', '3', '3', n, K, a_e, b_e, a_0, b_0, sigma_psi, sigma_prop1, sigma_prop2, nt, N_1, N_burn - N_1, order_penalty, lambda_penalty]
P = ['含义', '版本', '统计建模样条阶数（次数+1）', '时变参数样条阶数（次数+1）', '观测值个数', '基函数个数', '', '', '', '', '', '第一阶段采样sigma（Burn-in stage）', '第二阶段采样sigma（Steady stage）', '', 'Burn-in 阶段采样次数', 'Steady阶段采样次数', '惩罚差分矩阵阶数', '惩罚系数']
for i in range(1,19):
    ws['N%d'%i] = N[i-1]
    ws['O%d'%i] = O[i-1]
    ws['P%d'%i] = P[i-1]

def setCellStyle(st):
     """
     设置单元格格式
     params: st:就是第一步创建的工作表
     """

     # 边框
     border = Border(
         left=Side(border_style='thin', color='000000'),
         right=Side(border_style='thin', color='000000'),
         top=Side(border_style='thin', color='000000'),
         bottom=Side(border_style='thin', color='000000'),
     )

     # 对齐
     alignment = Alignment(
         horizontal='center',
         vertical='center',
         text_rotation=0,
         indent=0
     )

     # 字体     
     font = Font(
         name='Times New Roman',
         size=11,
         bold=False,
         italic=False,
         strike=False,
         color='000000'
     )

     for row, row_ind in zip(st.iter_rows(), range(1, st.max_row + 1)):
         for cell in row:
         	# 设置边框
            st[cell.coordinate].border = border
     		# 设置居中对齐
            st[cell.coordinate].alignment = alignment
            # 行高40
            st.row_dimensions[row_ind].height = 14
            # 设置字体
            st[cell.coordinate].font = font

      # 设置列宽         
     st.column_dimensions['A'].width = 6
     st.column_dimensions['B'].width = 14
     st.column_dimensions['C'].width = 14
     st.column_dimensions['D'].width = 14
     st.column_dimensions['E'].width = 14
     st.column_dimensions['F'].width = 14
     st.column_dimensions['G'].width = 14
  
     st.column_dimensions['I'].width = 14
     st.column_dimensions['J'].width = 14
     st.column_dimensions['k'].width = 14
     st.column_dimensions['L'].width = 14
     st.column_dimensions['N'].width = 14
     st.column_dimensions['O'].width = 35
     st.column_dimensions['P'].width = 35
     
     
     st.column_dimensions['H'].width = 2
     st.column_dimensions['M'].width = 2

setCellStyle(ws)
wb.save(dest_filename)
wb.close()

